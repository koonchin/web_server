import datetime
from io import BytesIO
import json,random,string
import re
import time
import numpy as np
from sqlalchemy import create_engine
import mysql.connector
import pandas as pd
import requests
import sqlalchemy
from pandas.api.types import CategoricalDtype
from PIL import Image, ImageDraw, ImageFont,ImageOps
from requests_toolbelt import MultipartEncoder
from sqlalchemy.sql.schema import MetaData
from app import settings
from server import cancelLoadpath,runLoadpath
from source import *
from collections import Counter
import xlsxwriter

class EMS:
    def __init__(self,token) -> None:
        self.main_token = token
    def GetTokenApi(self):
        endpoint = "https://trackapi.thailandpost.co.th/post/api/v1/authenticate/token"
        headers = {
            "Authorization": f"Token {self.main_token}",
            "Content-Type": "application/json"
        }
        # Make the POST request
        response = requests.post(endpoint, headers=headers)

        # Check the response
        if response.status_code == 200:
            # Successful response
            result = response.json()
            self.token = result['token']
        else:
            # Handle the error
            print(f"Error: {response.status_code} - {response.text}")

    def GetTokenWebhook(self):
        endpoint = "https://trackwebhook.thailandpost.co.th/post/api/v1/authenticate/token"
        headers = {
            "Authorization": f"Token {self.main_token}",
            "Content-Type": "application/json"
        }
        # Make the POST request
        response = requests.post(endpoint, headers=headers)

        # Check the response
        if response.status_code == 200:
            # Successful response
            result = response.json()
            self.token_webhook = result['token']
            print(self.token_webhook)
        else:
            # Handle the error
            print(f"Error: {response.status_code} - {response.text}")

    def GetItemsbyBarcode(self,barcode:list):
        self.GetToken()
        header = {'Authorization':"Token " +self.token,
                    'Content-Type':'application/json'}
        url = 'https://trackapi.thailandpost.co.th/post/api/v1/track'
        body = {
                "status": "all",
                "language": "TH",
                "barcode": barcode
                }
        res = requests.post(url=url,headers=header,json=body)
        print(res.text)

    def SubscribeByReceipt(self,receipt):
        self.GetTokenWebhook()
        header = {'Authorization':"Token " +self.token_webhook,
                    'Content-Type':'application/json'}
        url = 'https://trackwebhook.thailandpost.co.th/post/api/v1/hook'
        params = {
                "status": "all",
                "language": "TH",
                "barcode": receipt,
                "req_previous_status": "true",
                }

        res = requests.post(url=url,headers=header,json=params)
        print(res.text)

ems = EMS('ZhNQRIIFX+KZTbAeXTIxQ*TcK&B%G0BaFwXfY:YeUlR$UJBGH_E-F7N6CXK-GoTASaO6JERiAXDCFgBgF6HrChJ*JTQxG5YuQOSR')

class DB:

    def connect(self,dbname=databasedb):
        c = mysql.connector.connect(
            host=hostdb,
            user=userdb,
            password=passworddb,
            database=dbname,
            port='3306'
        )
        return c
    def query_with_image(self,query,args):
        c = self.connect()
        cursor = c.cursor(buffered=True)
        result = cursor.execute(
        query, args)
    # Committing the data
        c.commit()
    def query(self, task):
        try:
            c = self.connect()
            mycursor = c.cursor(buffered=True)
            mycursor.execute(task)
        except:
            c = self.connect()
            mycursor = c.cursor(buffered=True)
            mycursor.execute(task)
        return mycursor
    
    def check(self,task,db='muslin'):
        try:
            c = self.connect(dbname=db)
            mycursor = c.cursor(buffered=True)
            mycursor.execute(task)
        except:
            c = self.connect(dbname=db)
            mycursor = c.cursor(buffered=True)
            mycursor.execute(task)
        
        return mycursor.fetchall()

    def query_custom(self, task,db):
        try:
            c = self.connect(dbname=db)
            mycursor = c.cursor(buffered=True)
            mycursor.execute(task)
        except:
            c = self.connect(dbname=db)
            mycursor = c.cursor(buffered=True)
            mycursor.execute(task)
        return mycursor

    def query_commit(self, task):
        try:
            c = self.connect()
            mycursor = c.cursor(buffered=True)
            mycursor.execute(task)
            c.commit()
        except:
            c = self.connect()
            mycursor = c.cursor(buffered=True)
            mycursor.execute(task)
            c.commit()
        return mycursor

    def query_commit_many(self, task,rows):
        try:
            c = self.connect()
            mycursor = c.cursor(buffered=True)
            mycursor.executemany(task,rows)
            c.commit()
        except:
            c = self.connect()
            mycursor = c.cursor(buffered=True)
            mycursor.executemany(task,rows)
            c.commit()
        return mycursor

    def callproc(self, procname: str, task_db, task_db2):
        arg = [task_db, task_db2]
        try:
            c = self.connect()
            mycursor = c.cursor(buffered=True)
            mycursor.callproc(procname, arg)
            c.commit()
        except:
            c = self.connect()
            mycursor = c.cursor(buffered=True)
            mycursor.callproc(procname, arg)
            c.commit()
    
    def create_table(self,dbname):
        task = f'''create table if not exists {dbname} (  `sku` longtext NOT NULL,
                    `descript` longtext,
                    `amount` int DEFAULT NULL )'''
        db.query_commit(task)
        # 

    def insert_into_duplicate(self,dbname,data,amount):
        task = f"""
        INSERT INTO {dbname}(sku, descript, amount)
        VALUES ({data})
        ON DUPLICATE KEY UPDATE amount = {amount};  
        """
        # db.query_commit(task)
db = DB()

def add(sku, user):
    dep = get_role(user,"department")
    task_db = f'insert into {dep}.live_room values ("{sku}",NULL)'
    db.query_commit(task_db)
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    payload_list = []
    payload_list.append({"sku":sku,"stock":1,"cost":0})
    payload = {}
    payload['stocks'] = payload_list
    web.decrease_stock(payload,'W0002')
    db.query_commit(f"update {dep}.stock_main set amount = amount - 1 where sku = '{sku}'")
    db.query_commit(f"insert into {dep}.log values ('{user}','เพิ่มเข้าสต็อกห้องไลฟ์ รหัส {sku}',now())")
        
def delete(sku, user):
    dep = get_role(user,"department")
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    payload_list = []
    payload_list.append({"sku":sku,"stock":1,"cost":0})
    payload = {}
    payload['stocks'] = payload_list
    web.increase_stock(payload,'W0001')
    task_db = f'delete from {dep}.live_room where sku = "{sku}"'
    db.query_commit(task_db)

    db.query_commit(f"insert into {dep}.log values ('{user}','ลบสต็อกห้องไลฟ์ รหัส {sku}',now())")
    # send_line(f"รหัส {sku} เอาออกโดย {user}")
    if '-' in sku:
        sku = sku.split('-')[0]
    task_db = f"select stock_main.size from {dep}.stock_main\
    inner join {dep}.stock on {dep}.stock_main.sku = {dep}.stock.sku\
    where stock_main.amount + stock.amount > 0 and {dep}.stock_main.sku like '%{sku}%'\
    ORDER BY FIELD(stock_main.size, 'XXS', 'XS', 'S', 'M', 'L', 'F', 'XL', '2XL','3XL'), {dep}.stock_main.size"
    mycursor = db.query(task_db)
    myresult = list(mycursor.fetchall())
    if not len(myresult) == 0:
        return myresult[0][0]
    else:
        return False

def get_mysql_path():
    mydb = mysql.connector.connect(
        host=hostdb,
        user=userdb,
        password=passworddb,
        database=databasedb
    )
    mycursor = mydb.cursor()
    mycursor.execute('SHOW VARIABLES LIKE "secure_file_priv";')
    return mycursor.fetchone()[1]

def convert_url_to_bytes(url):
    try:
        response = requests.get(url, stream=True)
        data = BytesIO(response.content)
        return data.getvalue()
    except:
        return None

def get_api_register(department, api):
    """
    select {column} from store_api where department
    """
    mycursor = db.query(
        f"select {api} from store_api where department = '{department}'")
    try:
        return list(mycursor.fetchall())[0][0]
    except:
        return None

def get_role(req, role):
    """
        get {columns} from register_employee
    """

    if not type(req) == str:
        req = str(req.user)
    mycursor = db.query(
        f"select {role} from register_employee where user = '{req}'")
    try:
        return str(list(mycursor.fetchall())[0][0])
    except:
        return None

def send_line(msg):
    url = 'https://notify-api.line.me/api/notify'
    token = 'qGjBCv5oFvqb9IKNM05puRYhwDz0X3vAr6bs9y1Kvp8'
    headers = {'content-type': 'application/x-www-form-urlencoded',
               'Authorization': 'Bearer '+token}
    r = requests.post(url, headers=headers, data={'message': msg})

def send_line_webhook(msg):
    url = 'https://notify-api.line.me/api/notify'
    token = 'pd4KTq3uJa19WWuCQ5RALzISFos9ZTtOVXUsvctBWlW'
    headers = {'content-type': 'application/x-www-form-urlencoded',
               'Authorization': 'Bearer '+token}
    r = requests.post(url, headers=headers, data={'message': msg})

def send_line_adjust(msg):
    url = 'https://notify-api.line.me/api/notify'
    token = 'AyfLcJAM8h5IdNuxGUPqVfcFuS8Lt2dK6jCzKrPLIcW'
    headers = {'content-type': 'application/x-www-form-urlencoded',
               'Authorization': 'Bearer '+token}
    r = requests.post(url, headers=headers, data={'message': msg})

def send_line_return(msg):
    url = 'https://notify-api.line.me/api/notify'
    token = 'MMHcMhda3sy1ilLdq5FkDOyNNyp6HubLYNcInlq8ULx'
    headers = {'content-type': 'application/x-www-form-urlencoded',
               'Authorization': 'Bearer '+token}
    r = requests.post(url, headers=headers, data={'message': msg})

def send_line_facecheck(msg):
    url = 'https://notify-api.line.me/api/notify'
    token = 'dchmei9xBU2PffUQyqQsRskl6Sp9uYPljSOcEPy2yJJ'
    headers = {'content-type': 'application/x-www-form-urlencoded',
               'Authorization': 'Bearer '+token}
    r = requests.post(url, headers=headers, data={'message': msg})

def send_line_facecheck_img(image,msg):
    url = 'https://notify-api.line.me/api/notify'
    token = 'dchmei9xBU2PffUQyqQsRskl6Sp9uYPljSOcEPy2yJJ'
    headers = {'Authorization': 'Bearer ' + token}
    img = {'imageFile': open(image, 'rb')}
    data = {'message': msg}
    session = requests.Session()
    session_post = session.post(url, headers=headers, files=img, data=data)
    print(session_post.text)
    
class Web():
    def __init__(self, apikey, apisecret, storename) -> None:
        self.apikey = apikey
        self.apisecret = apisecret
        self.storename = storename
    
    def get_purchase_order(self,number):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://open-api.zortout.com/v4/PurchaseOrder/GetPurchaseOrders'
        params = {'keyword':number}
        res = requests.get(url=url, headers=header,params=params)
        data = res.json()
        for i in data['list']:
            print(i['number'])

    def get_order_vrich(self,orderdate):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://open-api.zortout.com/v4/Order/GetOrders'
        params = {'keyword':"VRM-",'orderdateafter':orderdate}
        res = requests.get(url=url, headers=header,params=params)
        data = res.json()
        number = [i['number'] for i in data['list']]
        return number

    def get_minus_available(self,warehousecode):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        sku = []
        payload_list = []
        url = 'https://open-api.zortout.com/v4/Product/GetProducts'
        for page in range(1,20):
            params = {'warehousecode':warehousecode,'page':page}
            res = requests.get(url=url, headers=header,params=params)
            data = res.json()
            for i in data['list']:
                if int(i['availablestock']) < 0 and str(i['sku']).startswith('M'):
                    payload_list.append({'sku':i['sku'],'stock':0,'cost':0})
                if int(i['availablestock']) < 0 and warehousecode == 'W0002':
                    payload_list.append({'sku':i['sku'],'stock':0,'cost':0})
        self.update_available_stock_list(warehousecode,payload_list)

    def get_sku_with_list(self,warehousecode,sku_list):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        sku,amount = [],[]
        payload_list = []
        count = 0
        url = 'https://open-api.zortout.com/v4/Product/GetProducts'
        for page in range(1,50):
            params = {'warehousecode':warehousecode,'page':page}
            res = requests.get(url=url, headers=header,params=params)
            data = res.json()
            for i in data['list']:
                if '-' in i['sku']:
                    sku_code = i['sku'].split('-')[0]
                    if sku_code in sku_list:
                        sku.append(i['sku'])
                        amount.append(i['stock'])
                        payload_list.append({'sku':i['sku'],'stock':0,'cost':0})
                        count += 1
                if count > 400:
                    self.update_available_stock_list(warehousecode,payload_list)
                    count = 0
                    payload_list = []

        self.update_available_stock_list(warehousecode,payload_list)
        return sku,amount

    def get_all_sku_by_warehouse(self,warehousecode):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        sku = []
        url = 'https://open-api.zortout.com/v4/Product/GetProducts'
        for page in range(1,20):
            params = {'warehousecode':warehousecode,'inventorystatus':1,'page':page}
            res = requests.get(url=url, headers=header,params=params)
            data = res.json()
            for i in data['list']:
                sku.append(i['sku'])
        return sku

    def get_order_detail(self,id):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        sku = []
        url = 'https://open-api.zortout.com/v4/Order/GetOrderDetail'
        params = {'id':id}
        res = requests.get(url=url, headers=header,params=params)
        data = res.json()
        for i in data['list']:
            sku.append(i['sku'])
        return sku

    def get_order_and_sku_amount(self,number):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}',
            'numberlist': number
        }

        url = 'https://open-api.zortout.com/v4/Order/GetOrders'
        res = requests.get(url=url, headers=header)
        data = res.json()
        print(number)
        for i in data['list']:
            if i['warehousecode'] == 'W0002':
                for product in i['list']:
                    sku = product['sku']
                    amount = product['number']
                    payload_list = []
                    payload_list.append({"sku":sku,"stock":int(amount),"cost":0})
                    payload = {}
                    payload['stocks'] = payload_list
                    self.decrease_stock(payload,'W0001')
        # self.update_order_status(number)

    def get_order_for_check(self,number):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}',
            'numberlist': number
        }

        url = 'https://open-api.zortout.com/v4/Order/GetOrders'
        params = {'number':number}
        res = requests.get(url=url, headers=header,params=params)
        data = res.json()
        return data['count']

    def get_transfer(self):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        df = pd.read_excel(r'D:\Dropbox\MUSLIN PAJAMAS ALL IN FO\Stock\ของเข้าร้านใหญ่\กันยายน 66\8\2\31925152 po 2.xlsm')
        df = df.iloc[1: , :]
        df.dropna(subset = ['SKU.1'], inplace=True)
        sku = [df.loc[i,'SKU.1'] for i in df.index]
        url = 'https://open-api.zortout.com/v4/Transfer/GetTransfers'
        param = {"createdafter":"2023-09-08","transfertype":"Adjust","page":2}
        res = requests.get(url=url, headers=header,params=param)
        data = res.json()
        for i in data['list']:
            for detail in i['list']:
                if detail['sku'] in sku:
                    self.void_transfer(i['id'])

    def void_transfer(self,id):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        url = 'https://open-api.zortout.com/v4/Transfer/VoidTransfer'
        param = {"id":id}
        res = requests.post(url=url, headers=header,params= param)
        print(res.text)
        
    def decrease_stock(self,payload,warehousecode):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        params = {'warehousecode':warehousecode}
        url = 'https://open-api.zortout.com/v4/Product/DecreaseProductStockList'

        response = requests.post(url=url, params=params,headers=header, json=payload)
        print(response.text)

    def increase_stock(self,payload,warehousecode):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        params = {'warehousecode':warehousecode}
        url = 'https://open-api.zortout.com/v4/Product/IncreaseProductStockList'

        response = requests.post(url=url, params=params,headers=header, json=payload)
        print(response.text)
        print(response.status_code)

    def update_available_stock_list(self,warehousecode,data):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        params = {'warehousecode':warehousecode}
        url = 'https://open-api.zortout.com/v4/Product/UpdateProductAvailableStockList'
        payload = {"stocks":data}
        response = requests.post(url=url, params=params,headers=header, json=payload)
        print(response.text)

    def transfer_all_amount_with_condition(self,condition_number:int,fromwarehousecode,towarehousecode,name):
        today = datetime.date.today()

        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        url = 'https://open-api.zortout.com/v4/Product/GetProducts'
        for page in range(1,30):
            params = {'warehousecode':fromwarehousecode,'limit':500,'page':page,'inventorystatus':1}
            response = requests.get(url=url, params=params,headers=header)
            data = response.json()
            count,count_post = 0,1
            payload_list = []
            for i in data['list']:
                sku,amount = i['sku'],int(i['availablestock'])
                if amount > condition_number and sku.lower() != 'fee':
                    payload_list.append({'sku':sku,'name':sku,'number':amount - condition_number})
                    count += 1
                if count > 450:
                    self.post_transfer(fromwarehousecode,towarehousecode,f'{name} {today} {page}/{count_post} 2',payload_list)
                    count_post += 1
                    payload_list = []
                    count = 0
            if len(payload_list) > 0:
                self.post_transfer(fromwarehousecode,towarehousecode,f'{name} {today} {page}/{count_post} last 2',payload_list)

    def transfer_all_amount_with_condition_by_idsell(self,idsell,fromwarehousecode,towarehousecode):
        today = datetime.date.today()

        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        print(idsell)
        url = 'https://open-api.zortout.com/v4/Product/GetProducts'
        for page in range(1,30):
            params = {'warehousecode':fromwarehousecode,'limit':500,'page':page,'inventorystatus':1}
            response = requests.get(url=url, params=params,headers=header)
            data = response.json()
            count,count_post = 0,1
            payload_list = []
            for i in data['list']:
                sku,amount = i['sku'],int(i['availablestock'])
                if amount > 0:
                    if sku.lower() != 'fee':
                        if sku.lower() != 'edit':
                            if get_idsell(sku) in idsell:
                                print(sku)
                                payload_list.append({'sku':sku,'name':sku,'number':amount})
                                count += 1
                if count > 450:
                    self.post_transfer(fromwarehousecode,towarehousecode,f'ดึง2 {today} {page}/{count_post} 2',payload_list)
                    count_post += 1
                    payload_list = []
                    count = 0
            if len(payload_list) > 0:
                self.post_transfer(fromwarehousecode,towarehousecode,f'ดึง2 {today} {page}/{count_post} last 2',payload_list)

    def set_zero_live_warehouse(self,warehousecode):
        today = datetime.date.today()

        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        url = 'https://open-api.zortout.com/v4/Product/GetProducts'
        for page in range(1,10):
            params = {'warehousecode':warehousecode,'limit':500,'page':page,'inventorystatus':1}

            response = requests.get(url=url, params=params,headers=header)
            data = response.json()
            count,count_post = 0,1
            payload_list = []
            for i in data['list']:
                sku = i['sku']
                count += 1
                payload_list.append({'sku':sku,'stock':0,'cost':0})
                if count > 450:
                    self.update_available_stock_list(warehousecode,payload_list)
                    count_post += 1
                    payload_list = []
                    count = 0
            if len(payload_list) > 0:
                self.update_available_stock_list(warehousecode,payload_list)

    def post_transfer(self,fromwarehouse,towarehouse,name_transfer,payload):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        param = {'uniquenumber':name_transfer}

        url = 'https://open-api.zortout.com/v4/Transfer/AddTransfer'
        data = {'number': name_transfer,'fromwarehousecode':fromwarehouse,'status':'Success','towarehousecode':towarehouse,'list':payload,"transferdate":f"{datetime.date.today()}"}
        print(data)
        # Convert the data dictionary to JSON
        json_data = json.dumps(data)
        res = requests.post(url=url, headers=header,data=json_data,params=param)
        print(res.text)
        print(res.status_code)

    #  upload trackingno today to database
    def post_order(self,data):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "ADDORDER", 'version': '3'}
        
        res = requests.post(url=url, headers=header, params=payload, json=data)
        print(res.status_code)

    def add_location(self,dep,sku,location):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        param = {"id":self.getId(sku)}
        task = f"select descript from stock_main where sku = '{sku}'"
        name = db.query_custom(task,dep)
        name = name.fetchall()[0][0]
        name = re.sub(r'\[.*?\]', '', name)
        
        url = 'https://open-api.zortout.com/v4/Product/UpdateProduct'
        descript = f'[{location}] {name}'
        data = {'name':descript}
        # Convert the data dictionary to JSON
        json_data = json.dumps(data)

        res = requests.post(url=url, headers=header, params=param, json=data)
        db.query_commit(f"update {dep}.stock_main set descript = '{descript}' where sku = '{sku}'")
        print(f"update {dep}.stock_main set descript = '{descript}' where sku = '{sku}'")

    def update_name(self,dep,sku,replace_word,replace_by_word):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        param = {"id":self.getId(sku)}
        task = f"select descript from stock_main where sku = '{sku}'"
        name = db.query_custom(task,dep)
        name = name.fetchall()[0][0]
        
        url = 'https://open-api.zortout.com/v4/Product/UpdateProduct'
        descript = name.replace(replace_word,replace_by_word)
        data = {'name':descript}
        # Convert the data dictionary to JSON
        json_data = json.dumps(data)

        res = requests.post(url=url, headers=header, params=param, json=data)

    def update_content_tag_to_iamge(self):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        url = 'https://open-api.zortout.com/v4/Product/GetProducts'
        for page in range(1,10):
            param = {'warehousecode':'W0001','page':page}
            res = requests.get(url=url, headers=header, params=param)
            if res.status_code == 200:
                data = res.json()
                for i in data['list']:
                    sku = i['sku']
                    barcode = i['barcode']
                    if barcode != sku:
                        self.updatebarcode(sku)
            else:
                print(f"Request failed with status code: {res.status_code}")

    def get_tag_by_sku(self,sku):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}',
            'skulist': sku
        }
        url = 'https://open-api.zortout.com/v4/Product/GetProducts'
        param = {'warehousecode':'W0002'}
        res = requests.get(url=url, headers=header, params=param)
        if res.status_code == 200:
            data = res.json()
            tag = data['list'][0]['tag']
            return tag
        else:
            return 'error',sku

    def get_not_qc_tag(self):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}',
        }
        url = 'https://open-api.zortout.com/v4/Product/GetProducts'
        for page in range(1,100):
            param = {'warehousecode':'W0001','page':page}
            res = requests.get(url=url, headers=header, params=param)
            data = res.json()
            for i in data['list']:
                tag = i['tag']
                if tag:
                    if 'สินค้ายังไม่ QC' in tag:
                        print(i['sku'],tag.remove('สินค้ายังไม่ QC'))
                        self.add_tag(i['id'],tag)
    
    def send_sales_report(self,dep,date):

        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://open-api.zortout.com/v4/Order/GetOrders'
        # datetime.datetime.now().strftime('%Y-%m-%d')
        if not date:
            date = (datetime.datetime.now() - datetime.timedelta(days = 1)).strftime('%Y-%m-%d')
        sales_channels_list = ['tiktok','lazada','vrich','shopee muslinpajamas','war','pare','shopee','facebook','line','ig','minny']
        sales_channels_dict = {}
        for i in sales_channels_list:
            sales_channels_dict[i] = []
        for page in range(1,4):
            params = {'method': "GETORDERS", 'version': '3','orderdateafter':date,'orderdatebefore':date,'page':page}
            res = requests.get(url=url, headers=header, params=params)
            print(res.content)
            data = res.json()
            for customer in data['list']:
                if customer['status'].lower() != 'voided':
                    if customer['saleschannel'].lower() in sales_channels_list:
                        sales_channels_dict[customer['saleschannel'].lower()].append(round(float(customer['amount']),2))
                
        if dep == 'muslin':
            sum_all = int(sum(sales_channels_dict['tiktok'])) +int(sum(sales_channels_dict['shopee muslinpajamas'])) + int(sum(sales_channels_dict['lazada'])) + int(sum(sales_channels_dict['war'])) + int(sum(sales_channels_dict['pare'])) + int(sum(sales_channels_dict['vrich']))
            TIKTOK_SUMALL,SHOPEE_SUMALL,LAZADA_SUMALL,ZORT_SUMALL,VRICH_SUMALL = format(int(sum(sales_channels_dict['tiktok'])),','),format(int(sum(sales_channels_dict['shopee muslinpajamas'])),','),format(int(sum(sales_channels_dict['lazada'])),','),format(sum(sales_channels_dict['war']) + sum(sales_channels_dict['pare']),','),format(int(sum(sales_channels_dict['vrich'])),',')
        else:
            sum_all = int(sum(sales_channels_dict['tiktok'])) +int(sum(sales_channels_dict['shopee'])) + int(sum(sales_channels_dict['lazada'])) + int(sum(sales_channels_dict['minny'])) +int(sum(sales_channels_dict['facebook'])) +int(sum(sales_channels_dict['line'])) + int(sum(sales_channels_dict['ig'])) +int(sum(sales_channels_dict['vrich']))
            TIKTOK_SUMALL,SHOPEE_SUMALL,LAZADA_SUMALL,ZORT_SUMALL,VRICH_SUMALL = format(int(sum(sales_channels_dict['tiktok'])),','),format(int(sum(sales_channels_dict['shopee'])),','),format(int(sum(sales_channels_dict['lazada'])),','),format(sum(sales_channels_dict['minny']) + sum(sales_channels_dict['facebook']) + sum(sales_channels_dict['ig']) + sum(sales_channels_dict['line']) ,','),format(int(sum(sales_channels_dict['vrich'])),',')
        
        result =  f"""คีย์มือ : {ZORT_SUMALL}
Shopee : {SHOPEE_SUMALL}
TikTok : {TIKTOK_SUMALL}
Lazada : {LAZADA_SUMALL}
Vrich : {VRICH_SUMALL}
ยอดรวม : {format(int(sum_all),',')}"""
        return result
    def send_sales_report_monthly(self,dep):

        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://open-api.zortout.com/v4/Order/GetOrders'
        # datetime.datetime.now().strftime('%Y-%m-%d')
        date = datetime.datetime.today().replace(day=1).strftime('%Y-%m-%d')
        sales_channels_list = ['tiktok','lazada','vrich','shopee muslinpajamas','war','pare','shopee','facebook','line','ig']
        sales_channels_dict = {}
        for i in sales_channels_list:
            sales_channels_dict[i] = []
        for page in range(1,20):
            params = {'method': "GETORDERS", 'version': '3','orderdateafter':date,'page':page}
            res = requests.get(url=url, headers=header, params=params)
            data = res.json()
            for customer in data['list']:
                if customer['status'].lower() != 'voided':
                    if customer['saleschannel'].lower() in sales_channels_list:
                        sales_channels_dict[customer['saleschannel'].lower()].append(round(float(customer['amount']),2))
                
        if dep == 'muslin':
            sum_all = int(sum(sales_channels_dict['tiktok'])) +int(sum(sales_channels_dict['shopee muslinpajamas'])) + int(sum(sales_channels_dict['lazada'])) + int(sum(sales_channels_dict['war'])) + int(sum(sales_channels_dict['pare'])) + int(sum(sales_channels_dict['vrich']))
            TIKTOK_SUMALL,SHOPEE_SUMALL,LAZADA_SUMALL,ZORT_SUMALL,VRICH_SUMALL = format(int(sum(sales_channels_dict['tiktok'])),','),format(int(sum(sales_channels_dict['shopee muslinpajamas'])),','),format(int(sum(sales_channels_dict['lazada'])),','),format(sum(sales_channels_dict['war']) + sum(sales_channels_dict['pare']),','),format(int(sum(sales_channels_dict['vrich'])),',')
        else:
            sum_all = int(sum(sales_channels_dict['tiktok'])) +int(sum(sales_channels_dict['shopee'])) + int(sum(sales_channels_dict['lazada'])) + int(sum(sales_channels_dict['facebook'])) + int(sum(sales_channels_dict['line'])) + int(sum(sales_channels_dict['ig']))  + int(sum(sales_channels_dict['vrich']))
            TIKTOK_SUMALL,SHOPEE_SUMALL,LAZADA_SUMALL,ZORT_SUMALL,VRICH_SUMALL = format(int(sum(sales_channels_dict['tiktok'])),','),format(int(sum(sales_channels_dict['shopee'])),','),format(int(sum(sales_channels_dict['lazada'])),','),format(sum(sales_channels_dict['facebook']) + sum(sales_channels_dict['ig']) + sum(sales_channels_dict['line']),','),format(int(sum(sales_channels_dict['vrich'])),',')
        
        result =  f"""คีย์มือ : {ZORT_SUMALL}
Shopee : {SHOPEE_SUMALL}
TikTok : {TIKTOK_SUMALL}
Lazada : {LAZADA_SUMALL}
Vrich : {VRICH_SUMALL}
ยอดรวม : {format(int(sum_all),',')}"""
        return result

    def addImage(self, sku, imagepath):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "UPDATEPRODUCTIMAGE", 'version': '3',
                   'id': self.getId(sku)}

        data = {'file': open(imagepath, 'rb')}
        res = requests.post(url=url, headers=header,
                            params=payload, files=data)
        print('here',res.text)

    def add_image_to_order(self, number,img_url):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://open-api.zortout.com/v4/Order/AddOrderFile'

        payload = {'number': number}

        # Fetch the file content
        with requests.get(img_url, stream=True) as response:
            # Check if the request was successful (status code 200)
            if response.status_code == 200:
                # Set up the file parameter for your API
                files = {"file": (f"{number}.jpg", response.raw, 'image/jpeg')}

                # Make a POST request to your API endpoint with the file parameter
                res = requests.post(url=url, headers=header, params=payload, files=files)
                print('here', res.text)
            else:
                print(f"Failed to fetch the file. Status code: {response.status_code}")

    def getId(self,sku):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "GETPRODUCTS", 'version': '3',
                   'searchsku': sku}

        res = requests.get(url=url, headers=header,
                            params=payload)
        print(res.status_code)
        print(res.content)
        data = res.json()
        return data['list'][0]['id']

    def getAvailable(self,sku):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "GETPRODUCTS", 'version': '3',
                   'searchsku': sku}

        res = requests.get(url=url, headers=header,
                            params=payload)
        data = res.json()
        return data['list'][0]['availablestock']

    def update_payment(self,number,paymentmethod,paymentamount):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "UPDATEORDERPAYMENT", 'version': '3','number':number,'paymentmethod':paymentmethod,'paymentamount':paymentamount}
        
        res = requests.post(url=url, headers=header, params=payload)
        print(res.status_code)

    def update_order_status(self,number):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "UPDATEORDERSTATUS", 'version': '3','number':number,'status':1}
        
        res = requests.post(url=url, headers=header, params=payload)
        print(number,res.status_code)
    
    def add_tracking(self,number,tracking):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "EDITORDERINFO", 'version': '3','number':number}
        data = {'trackingno':tracking}
        res = requests.post(url=url, headers=header, params=payload,json=data)
        res = json.loads(res.content.decode('utf-8'))
        return int(res['resCode']) == 200

    def VOIDORDERPAYMENT(self,number,paymentid):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "VOIDORDERPAYMENT", 'version': '3','number':number,'paymentid':paymentid}
        
        res = requests.post(url=url, headers=header, params=payload)
        print(number,res.content)
        
    def update_fee(self):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'
        day = datetime.datetime.now() - datetime.timedelta(days = 8)
        day = day.strftime('%Y-%m-%d')
        for i in range(20):
            payload = {'method': "GETORDERS", 'version': '3','status':"0",'orderdatebefore':day,'page':i}
            res = requests.get(url=url, headers=header, params=payload)
            data = res.json()
            for customer in data['list']:
                for info in customer['list']:
                    if 'FEE' in info['sku'].upper():
                        self.update_order_status(customer['number'])
    
    def update_Payment(self):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'
        day = datetime.datetime.now() - datetime.timedelta(days = 60)
        day = day.strftime('%Y-%m-%d')
        for i in range(20):
            payload = {'method': "GETORDERS", 'version': '3','paymentstatus':"0,3,4",'orderdateafter':day,'page':i}
            res = requests.get(url=url, headers=header, params=payload)
            data = res.json()
            for customer in data['list']:
                if 'COD' not in customer['shippingchannel']:
                    for info in customer['payments']:
                            self.VOIDORDERPAYMENT(customer['number'],info['id'])
                        # print(customer['number'])
                    self.update_payment(customer['number'],'โอนผ่านธนาคาร',int(customer['amount']))

    def updateproduct(self, sku, sellprice):
        header = {
            'storename': self.storename,
            'apikey': self.apikey,
            'apisecret': self.apisecret
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "UPDATEPRODUCT", 'version': '3'}
        data = {'sku':sku,'barcode':sku,'sellprice': sellprice}

        # Convert the data dictionary to JSON
        json_data = json.dumps(data)

        # Make the POST request
        res = requests.post(url=url, headers=header, params=payload, data=json_data)

        # Print the response
        print(res.text)
        print(res.url)

    def updatebarcode(self,sku):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "UPDATEPRODUCT", 'version': '3'}
        data = {'sku':sku,'barcode':sku}
        res = requests.post(url=url, headers=header, params=payload, json=data)
        print(sku,res.status_code)
        
    def updateproductName(self,sku,name):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "UPDATEPRODUCT", 'version': '3'}
        data = {'sku':sku,'name':name}
        res = requests.post(url=url, headers=header, params=payload, json=data)
        print(name,res.status_code)

    def get_product_for_price(self):
        
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        dep = 'muslin'
        url = 'https://open-api.zortout.com/v4/Product/GetProducts'
        taskdb_stock = ''
        for pages in range(30):
            payload = {'warehousecode': "W0001",'activestatus':1, 'page': pages,'limit':2000}
            res = requests.get(url=url, headers=header, params=payload)
            data = res.json()
            if len(data['list']) > 0:
                for i in range(len(data['list'])):
                    if data['list'][i]['sellprice'] == '0':
                        try:
                            self.updateproduct(data['list'][i]['id'],tran_price('muslin',data['list'][i]['sku']))
                            taskdb_stock += f"WHEN '{data['list'][i]['sku']}' THEN {int(tran_price('muslin',data['list'][i]['sku']))}\n"   
                        except:
                            print(data['list'][i]['sku'],'error')
            task_final = f'''
            UPDATE {dep}.cost
            SET price
            = CASE sku
            {taskdb_stock}
            ELSE price
            END;
            '''
            db.query_commit(task_final)

    def post_purchase_order(self,data):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "ADDPURCHASEORDER", 'version': '3'}

        res = requests.post(url=url, headers=header, params=payload, json=data)
        print(res.status_code)

    def get_track_chino(self,dep):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'
        day = datetime.datetime.now() - datetime.timedelta(days = 1)
        payload = {'method': "GETORDERS", 'version': '3','orderdateafter':day.strftime('%Y-%m-%d')}

        res = requests.get(url=url, headers=header, params=payload)
        data = res.json()
        count = 0 
        sku_list = []
        for i in range(len(data['list'])):
            if data['list'][i]['status'] == 'Success':
                if data['list'][i]['amount'] > 0:
                            for sku in data['list'][i]['list']:
                                sku_list.append(sku['sku'])
        return sku_list
    
    def get_track(self,dep):
        def check(idorder):
            try:
                result = db.query(f"select * from {dep}.deli_zort where idorder = {idorder}")
                result = list(result.fetchall())
                return result
            except Exception as e:
                print(e)
            return False

        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'
        day = datetime.datetime.now() - datetime.timedelta(days = 40)
        payload = {'method': "GETORDERS", 'version': '3','orderdateafter':day.strftime('%Y-%m-%d')}

        res = requests.get(url=url, headers=header, params=payload)
        data = res.json()
        count = 0 
        for i in range(len(data['list'])):
            if data['list'][i]['status'] == 'Waiting':
                trackingno = data['list'][i]['trackingno']
                IDorder = data['list'][i]['id']
                if check(IDorder):
                    db.query_commit(f"update {dep}.deli_zort set trackingno ='{trackingno}',printed = 1,printedtime = now() ,status = '{data['list'][i]['status']}'\
                        where idorder = {IDorder}")
                    count += 1
        return count
    
    def get_telephone(self,pages):
        def check(id,status,shipping,track,pay,number,cstname,orderdate,addr,tel):
            if id in idorder:
                return ""
                if not shipping:
                    task = f"update {dep}.deli_zort set status = '{status}',trackingno='{track}',paymentstatus='{pay}' ,tel='{tel}',addr = '{addr}'where idorder = {id};"
                else:
                    task = f"update {dep}.deli_zort set status = '{status}',trackingno = '{track}',paymentstatus='{pay}' ,tel='{tel}',addr = '{addr}'where idorder = {id};"
            else:
                task = f"""insert into {dep}.deli_zort
                values ('{id}','{number}','{status}','{cstname}','{orderdate}','',0,NULL,NULL,'{pay}',"{addr}",'{tel}');\n"""
                with open ("sql.txt","a",encoding="utf-8") as f:
                    f.write(task)


            # db.query_commit(task)
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}',
        }

        url = 'https://api.zortout.com/api.aspx'
        day = datetime.datetime.now() - datetime.timedelta(days = 90)
        # day1 = datetime.datetime.now() - datetime.timedelta(days = 2)
        payload = {'method': "GETORDERS", 'version': '3','orderdateafter':day.strftime('%Y-%m-%d'),'page':pages}
        # payload = {'method': "GETORDERS", 'version': '3','orderdateafter':"2024-03-31"}
        # payload = {'method': "GETORDERS", 'version': '3'}

        res = requests.get(url=url, headers=header, params=payload)
        data = res.json()
        # for i in range(len(data['list'])):
        #     if data['list'][i]['status'] == 'Pending':
        #         if data['list'][i]['paymentstatus'] == "Paid":
        #             IDorder = data['list'][i]['number']
    #             print(IDorder)
        for i in range(len(data['list'])):
            if data['list'][i]['amount'] > 0:
                if '"' in data['list'][i]['shippingaddress'] or "'" in data['list'][i]['shippingaddress']:
                    data['list'][i]['shippingaddress'] = data['list'][i]['shippingaddress'].replace('"','')
                    data['list'][i]['shippingaddress'] = data['list'][i]['shippingaddress'].replace("'",'')
                check(data['list'][i]['id'],data['list'][i]['status'],data['list'][i]['shippingdateString'],data['list'][i]['trackingno'],data['list'][i]['paymentstatus'],data['list'][i]['number'],data['list'][i]['customername'],data['list'][i]['orderdateString'],data['list'][i]['shippingaddress'],data['list'][i]['shippingphone'])
                for sku in data['list'][i]['list']:
                    if data['list'][i]['id'] not in idorder_ordermain:
                        task = f"insert into {dep}.order_main values ('{data['list'][i]['id']}','{sku['pricepernumber']}','{sku['sku']}','{data['list'][i]['amount']}','{sku['number']}');\n"
                        with open ("sql.txt","a",encoding="utf-8") as f:
                            f.write(task)
                        # db.query_commit(f"insert into {dep}.order_main values ('{data['list'][i]['id']}','{sku['pricepernumber']}','{sku['sku']}','{data['list'][i]['amount']}','{sku['number']}')")
                    
    def update_cost(self,pages):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        dep = 'maruay'
        url = 'https://api.zortout.com/api.aspx'
        payload = {'method': "GETPRODUCTS", 'version': '3','page':pages,'limit':1000,'warehousecode':"W0001"}

        res = requests.get(url=url, headers=header, params=payload)
        data = res.json()
        count = 0 
        for i in range(len(data['list'])):
            cost = data['list'][i]['purchaseprice']
            price = data['list'][i]['sellprice']
            sku = data['list'][i]['sku']
            
            task = f"""insert into {dep}.cost
            values ('{sku}',{cost},1,{price});\n"""
            with open ("sql.txt","a",encoding="utf-8") as f:
                f.write(task)

    def get_track_2(self,dep):
        result = db.query(f"select idorder from {dep}.deli_zort")
        result = list(result.fetchall())
        idorder = [i[0] for i in result]

        result = db.query(f"select idorder from {dep}.order_main")
        result = list(result.fetchall())
        idorder_ordermain = [i[0] for i in result]

        def check(id, status, shipping, track, pay, number, cstname, orderdate, addr, tel):
            if id in idorder:
                return None
            else:
                task = (f"('{id}','{number}','{status}','{cstname}','{orderdate}','',0,NULL,NULL,'{pay}',"
                        f"'{addr}','{tel}')")
                return task

        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}',
        }

        url = 'https://open-api.zortout.com/v4/Order/GetOrders'
        day = datetime.datetime.now() - datetime.timedelta(days=15)
        for pages in range(10):
            payload = {'method': "GETORDERS", 'version': '3', 'orderdateafter': day.strftime('%Y-%m-%d'), 'page': pages}

            try:
                res = requests.get(url=url, headers=header, params=payload)
                res.raise_for_status()  # Raise an exception if response status code is not 2xx
                data = res.json()
            except requests.exceptions.RequestException as e:
                print(f"An error occurred: {e}")
                # Handle the exception here, e.g. retry the request, log the error, etc.

            insert_deli_zort_tasks = []
            insert_order_main_tasks = []

            for i in range(len(data['list'])):
                if data['list'][i]['amount'] > 0:
                    if '"' in data['list'][i]['shippingaddress'] or "'" in data['list'][i]['shippingaddress']:
                        data['list'][i]['shippingaddress'] = data['list'][i]['shippingaddress'].replace('"', '')
                        data['list'][i]['shippingaddress'] = data['list'][i]['shippingaddress'].replace("'", '')
                    if '"' in data['list'][i]['customername'] or "'" in data['list'][i]['customername']:
                        data['list'][i]['customername'] = data['list'][i]['customername'].replace('"', '')
                        data['list'][i]['customername'] = data['list'][i]['customername'].replace("'", '')
                    check_task = check(data['list'][i]['id'], data['list'][i]['status'], data['list'][i]['shippingdateString'],
                                    data['list'][i]['trackingno'], data['list'][i]['paymentstatus'], data['list'][i]['number'],
                                    data['list'][i]['customername'], data['list'][i]['orderdateString'],
                                    data['list'][i]['shippingaddress'], data['list'][i]['shippingphone'])
                    if check_task:
                        insert_deli_zort_tasks.append(check_task)

                    for sku in data['list'][i]['list']:
                        if data['list'][i]['id'] not in idorder_ordermain:
                            task = f"('{data['list'][i]['id']}','{sku['pricepernumber']}','{sku['sku']}','{data['list'][i]['amount']}','{sku['number']}')"
                            insert_order_main_tasks.append(task)

            if insert_deli_zort_tasks:
                deli_zort_values = ',\n    '.join(insert_deli_zort_tasks)
                concat_deli_zort_tasks = f"INSERT INTO {dep}.deli_zort (idorder, number, status, customername, orderdate, trackingno, printed, printedtime, shippingtime, paymentstatus, addr, tel) VALUES\n    {deli_zort_values};\n"
                try:
                    db.query_commit(concat_deli_zort_tasks)
                except:
                    print(concat_deli_zort_tasks)
            if insert_order_main_tasks:
                order_main_values = ',\n    '.join(insert_order_main_tasks)
                concat_order_main_tasks = f"INSERT INTO {dep}.order_main (IDorder, price, sku, total, amount) VALUES\n    {order_main_values};\n"

                db.query_commit(concat_order_main_tasks)

    def get_track_success(self):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "GETORDERS", 'version': '3',"createdafter":"2021-11-01"}

        res = requests.get(url=url, headers=header, params=payload)
        data = res.json()

        for i in range(len(data['list'])):
            if data['list'][i]['status'] == 'Success':
                trackingno = data['list'][i]['trackingno']
                if trackingno:
                    IDorder = data['list'][i]['id']
                    date = data['list'][i]['createdatetimeString']
                    FBname = data['list'][i]['reference']
                    cstname = data['list'][i]['customername']
                    addr = data['list'][i]['customeraddress']
                    tel = data['list'][i]['customerphone']
                    amount = len(data['list'][i]['list'])
                    paid = data['list'][i]['paymentamount']
                    total = data['list'][i]['amount']
                    descript = [data['list'][i]['list'][i2]['sku'] for i2 in range(amount)]
                    price = [data['list'][i]['list'][i2]['totalprice'] for i2 in range(amount)]
                    status = 'packed'
                    result = db.query(f"select idorder from muslin.order_main where idorder = '{IDorder}'")
                    result = list(result.fetchall())
                    if not result:
                        now = time.strftime('%Y-%m-%d %H:%M:%S')
                        for i3 in range(amount):
                            print(IDorder)
                            try:
                                task =f"""
                                insert into muslin.order_main
                                values ('{IDorder}', '{date}', '{FBname}', '{cstname}', '{addr}', '{tel}', '{trackingno}', '{amount}', '{total}', '{paid}', "{now}", '{descript[i3]}', 'Zort', '{price[i3]}', '{status}')"""
                                db.query_commit(task)
                            except Exception as e:
                                print(e,IDorder)
                # for dt in range(len(data['list'][i]['list'])):
                #     sku = data['list'][i]['list'][dt]['sku']
                #     descript = data['list'][i]['list'][dt]['name']
    # get data from list of www.api.zort.com
    #  upload trackingno today to database
    def get_track_data(self,idorder,trackingNo):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "GETORDERDETAIL", 'version': '3','id':idorder}

        res = requests.get(url=url, headers=header, params=payload)
        data = res.json()
        date = data['createdatetimeString']
        FBname = data['reference']
        cstname = data['customername']
        addr = data['customeraddress']
        tel = data['customerphone']
        amount = len(data['list'])
        paid = data['paymentamount']
        total = data['amount']
        descript = [data['list'][i]['name'] for i in range(amount)]
        price = [data['list'][i]['totalprice'] for i in range(amount)]
        status = 'packed'
        return idorder, date, FBname, cstname, addr, tel, trackingNo, amount, total, paid, 'now()', descript, 'Zort', price, status
    # get data from list of www.api.zort.com

    def get(self, datatype, page=1):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "GETPRODUCTS", 'version': '3',
                   'page': page}
        res = requests.get(url=url, headers=header, params=payload)
        data = res.json()
        for i in range(len(data['list'])):
            if data['list'][i]['number'][:2] == 'SO' and data['list'][i]['status'] != 'Voided' and data['list'][i][''] != 'Excess Payment':
                number.append(data['list'][i]['number'])
                datalist.append(data['list'][i]['list'])
                status.append(data['list'][i]['status'])
        return number,status,datalist

    def get_order_shopee(self, page=1):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        task = """select idorder from muslin.deli_zort
        where shippingtime = '2021-12-17'"""
        result = db.query(task)
        result = list(result.fetchall())
        result = [i[0] for i in result]

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "GETORDERS", 'version': '3',
                   'page': page}
        res = requests.get(url=url, headers=header, params=payload)
        data = res.json()
        number = []
        for i in range(len(data['list'])):
            if data['list'][i]['id'] in result:
                number.append(data['list'][i]['list'])
        return number

    def post_zero(self, page=1,dep='muslin'):

        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "GETPRODUCTS", 'version': '3',
                    'page': page}

        res = requests.get(url=url, headers=header, params=payload)
        data = res.json()
        datadict = {}
        datadict['stocks'] = list()
        taskdb_stock = ''
        for i in range(len(data['list'])):
            if i % 200 == 0:
                self.postzero(datadict)
                datadict['stocks'] = list()
                print("{:.2f} %".format(i / len(data['list']) * 100))
            if int(data['list'][i]['stock']) != 0 and 'Q' not in data['list'][i]['sku']:
                datadict['stocks'].append({"sku": data['list'][i]['sku'], "stock": 0,'cost':0})
                # self.post("UPDATEPRODUCTSTOCKLIST",data['list'][i]['sku'],0)
            if int(data['list'][i]['availablestock']) < 0:
                taskdb_stock += f"WHEN '{data['list'][i]['sku']}' THEN {int(data['list'][i]['availablestock'])}\n"   
        self.postzero(datadict)
        return taskdb_stock

    def post_zero_path(self, path,page=1):
        df = pd.read_excel(path)
        sku = [df.loc[i,'sku'] for i in df.index]
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "GETPRODUCTS", 'version': '3',
                    'page': page}

        res = requests.get(url=url, headers=header, params=payload)
        data = res.json()
        datadict = {}
        datadict['stocks'] = list()
        taskdb_stock = ''
        for i in range(len(data['list'])):
            if i % 200 == 0:
                self.postzero(datadict)
                datadict['stocks'] = list()
                print("{:.2f} %".format(i / len(data['list']) * 100))
            if int(data['list'][i]['stock']) != 0 and data['list'][i]['sku'].split('-')[0] in sku:
                datadict['stocks'].append({"sku": data['list'][i]['sku'], "stock": 0,'cost':0})
                # self.post("UPDATEPRODUCTSTOCKLIST",data['list'][i]['sku'],0)
        self.postzero(datadict)
        return taskdb_stock

    def get_condition_2(self,  page=1):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "GETORDERS", "orderdatebefore":"2022-02-04",'version': '3',
                   'page': page}

        res = requests.get(url=url, headers=header, params=payload)
        data = res.json()
        datalist = []
        for i in range(len(data['list'])):
            if data['list'][i]['status'] != "Success":
                if  data['list'][i]['status'] != "Voided":
                    print(data['list'][i]['number'])
                    for sku in data['list'][i]['list']:
                       self.post("UPDATEPRODUCTAVAILABLESTOCKLIST",sku['sku'],int(sku['number']))
                    self.post_status_order(data['list'][i]["id"])
        return datalist

    def add_tag(self, id,tag):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = "https://open-api.zortout.com/v4/Product/UpdateProduct"
        param = {
            'id': id
        }
        
        # Encode the payload as UTF-8
        payload = {
            "tag": tag
        }
        
        response = requests.post(url, headers=header, params=param, json=payload)
        
        # Check the response status code and handle it accordingly
        if response.status_code == 200:
            print(response.text)
        else:
            print(f"Request failed with status code: {response.status_code}")

    def update_shipping_order(self,id):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = "https://open-api.zortout.com/v4/Order/UpdateOrderStatus"
        param = {
            'id':id,
            'status':1
        }
        response = requests.post( url, headers=header, params=param)
        if response.status_code == 200:
            return 'success' 
        else:
            return  'error'

    def update_order_by_track(self,barcode):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        url = "https://open-api.zortout.com/v4/Order/GetOrders"
        param = {
            'keyword':barcode,
        }
        response = requests.get( url, headers=header, params=param)
        data = response.json()
        result = []
        for i in range(len(data['list'])):
            res = self.update_shipping_order(data['list'][i]['id'])
            result.append(f"{data['list'][i]['number']} {res}")
        return result

    def get_tracking_no_by_order_number(self,ordernumber):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}',
            'numberlist': ordernumber
        }
        url = "https://open-api.zortout.com/v4/Order/GetOrders"
        response = requests.get( url, headers=header)
        data = response.json()
        if data['list']:
            trackingno = data['list'][0]['trackingno']
        else:
            trackingno = False
        return trackingno

    def update_order_shipping(self):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }
        url = "https://open-api.zortout.com/v4/Order/GetOrders"
        for page in range(1,10):
            param = {
                'orderdatebefore':'2023-10-01',
                'orderdateafter':'2023-08-01',
                'limit':2000,
                'keyword':'SO-2023',
                'page':page

            }
            response = requests.get( url, headers=header, params=param)
            data = response.json()
            result = []
            print(len(data['list']))
            for i in range(len(data['list'])):
                if data['list'][i]['status'].lower() == 'shipping':
                    res = self.update_shipping_order(data['list'][i]['id'])
                    print(f"{data['list'][i]['number']} {res}")
        return result
    
    def get_datatype_filter_by_sku(self,dep,page):
        task = f'select sku,cost from cost'
        result = db.query_custom(task,dep)
        result = list(result.fetchall())
        cost = {}
        for i in result:
            cost[i[0]] = int(i[1])

        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "GETPRODUCTS", 'version': '3','page':page}

        res = requests.get(url=url, headers=header, params=payload)
        data = res.json()
        for i in range(len(data["list"])):
            sku = data['list'][i]['sku']
            if 'Q' in sku:
                continue
            try:
                if cost[sku] == 0 and float(data['list'][i]['purchaseprice']) != 0:
                    print(sku)
                if cost[sku] != 0 and float(data['list'][i]['purchaseprice']) == 0:
                    print(sku)
            except Exception as E:
                print(E,sku)
    # post data to www.api.zort.com

    def post_status_order(self, id):

        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "UPDATEORDERSTATUS", 'version': '3', 'warehousecode': 'W0001',"id":int(id)}
        res = requests.post(url=url, headers=header, params=payload)
        print(res.status_code)

    def post(self, method, sku, amount,cost=0):

        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': method, 'version': '3', 'warehousecode': 'W0001'}
        data = {"stocks": [{"sku": f"{sku}", "stock": amount,'cost':cost}]}
        res = requests.post(url=url, headers=header, params=payload, json=data)
        # for i in range(len(data['list'])):
        #     print(data['list'][i]['customerphone'])

    def postzero(self,data):
        # example
        # data = {"stocks": [{"sku": "BA0002-S", "stock": 0,'cost':0},{"sku": "BA0002-L", "stock": 0,'cost':0},{"sku": "BA0002-M", "stock": 0,'cost':0}]}
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': 'UPDATEPRODUCTSTOCKLIST', 'version': '3', 'warehousecode': 'W0001'}
        res = requests.post(url=url, headers=header, params=payload, json=data)
        print(res.status_code)
        # print(json.loads(res.text)['resDesc'])

    def post_descript(self, method, sku, descript):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': method, 'version': '3', 'warehousecode': 'W0001'}
        data = {"sku": f'{sku}', 'description': f'{descript}'}
        res = requests.post(url=url, headers=header, params=payload, json=data)
        data = res.json()

    def update_track(self, idorder):
        header = {
            'storename': f'{self.storename}',
            'apikey': f'{self.apikey}',
            'apisecret': f'{self.apisecret}'
        }

        url = 'https://api.zortout.com/api.aspx'

        payload = {'method': "UPDATEORDERPAYMENT", 'version': '3', 'number': f"{idorder}",'paymentamount':0.1,"paymentmethod":"โอนผ่านธนาคาร"}
        res = requests.post(url=url, headers=header, params=payload)
        print(res.status_code)
        print(res.content)
        # for i in range(len(data['list'])):
        #     print(data['list'][i]['customerphone'])

def update_size(apikey,apisecret,storename):
    web =Web(apikey,apisecret,storename)
    df = pd.read_excel(r"C:\Users\Chino\Desktop\diff.xlsx")
    for i in df.index:
        if  not str(df.loc[i,"SIZE"]) == 'nan':
            web.post_descript("UPDATEPRODUCT",str(df.loc[i,'SKU']),str(df.loc[i,'SIZE']))
    
def export_excel_zortform(less, greate, path):
    df = pd.DataFrame(list(get_data(less, greate)), columns=[
        'รหัสสินค้า', 'ชื่อสินค้า', 'จำนวน', 'ทุน'])
    df['ชื่อคลัง/สาขา'] = 'คลังหลัก'
    today_date = datetime.date.today()
    today_date = today_date.strftime("%d/%m/%Y")
    df['วันที่ทำรายการ'] = ''
    df['วันที่จัดส่ง'] = ''
    df['หมายเลขรายการ'] = ''
    df['ชื่อคู่ค้า'] = ''
    df.at[0, 'หมายเลขรายการ'] = 'VR LIVE'+today_date
    df.at[0, 'ชื่อคู่ค้า'] = 'VR'
    df.at[0, 'วันที่ทำรายการ'] = today_date
    df.at[0, 'วันที่จัดส่ง'] = today_date

    df = df[['หมายเลขรายการ', 'ชื่อคู่ค้า', 'วันที่ทำรายการ', 'วันที่จัดส่ง',
             'ชื่อคลัง/สาขา', 'รหัสสินค้า', 'ชื่อสินค้า', 'จำนวน', 'ทุน']]
    df.to_excel(f'{path}', index=False)
    # result = list(export_excel(5,5))

def export_excel_vrichform_(less, greate, path):
    df = pd.DataFrame(list(get_data(less, greate)), columns=[
        'รหัสสินค้า', 'รายละเอียด', 'จำนวน', 'ต้นทุน'])

    def tran(var: str):
        result = var.split('0')[0]
        if len(result) == 2:
            firstchar = result[1]
        else:
            firstchar = result
        try:
            number = int(var.replace(result, '')[:4])
        except:
            print(var)
            return 'wrong'
        return firstchar+str(number)

    df['รหัสขาย'] = df['รหัสสินค้า'].apply(tran)
    df = df[df['รหัสขาย'] != 'wrong']
    df['หน่วย'] = ''
    df['ราคา'] = 790
    df['ค่าส่งเพิ่ม'] = 0
    df['หมายเหตุ'] = ''
    df['จำนวน'] = df['จำนวน'].apply(pd.to_numeric, errors='ignore')
    df['ต้นทุน'] = df['ต้นทุน'].apply(pd.to_numeric, errors='ignore')
    df = df[['รหัสสินค้า', 'รหัสขาย', 'รายละเอียด', 'หน่วย',
             'จำนวน', 'ราคา', 'ต้นทุน', 'ค่าส่งเพิ่ม', 'หมายเหตุ']]
    df.to_excel(f'{path}', index=False)

def send_message_facebook(recipient_id, message_text):
    params = {
        "access_token": 'EAACNi6PriH4BAHx71ZCuZADZCVWQ0v4ZABJt4IpEN36LCmdYRbYta5oVjcBel7bwiZBV2QU6EN1y0LywjNiHKYAvcOsWslUzJL8ZAuLZAVgP4GNRZCMCZBLw54alXAfsnwVxlHZCnQxY1Qkp8088HiMuBTeecFKrvqJSxlwsXvlH59iKCR9H9EVhFYjWV1yNt8BS4pO8b4CH781QP8ClSB7r33'
    }
    print(os.getcwd())
    data = {
        # encode nested json to avoid errors during multipart encoding process
        'recipient': json.dumps({
            'id': recipient_id
        }),
        # encode nested json to avoid errors during multipart encoding process
        'message': json.dumps({
            'attachment': {
                'type': 'image',
                'payload': {}
            }
        }),
        'filedata': (os.path.basename(message_text), open(message_text, 'rb'), 'image/png')
    }

    # multipart encode the entire payload
    multipart_data = MultipartEncoder(data)

    # multipart header from multipart_data
    multipart_header = {
        'Content-Type': multipart_data.content_type
    }
    r = requests.post("https://graph.facebook.com/v12.0/me/messages",
                      params=params, headers=multipart_header, data=multipart_data)
    if r.status_code != 200:
        print(r.status_code)
        print(r.text)

def get_photo_facebook(id):
    params = {
        "access_token": 'EAACNi6PriH4BAHx71ZCuZADZCVWQ0v4ZABJt4IpEN36LCmdYRbYta5oVjcBel7bwiZBV2QU6EN1y0LywjNiHKYAvcOsWslUzJL8ZAuLZAVgP4GNRZCMCZBLw54alXAfsnwVxlHZCnQxY1Qkp8088HiMuBTeecFKrvqJSxlwsXvlH59iKCR9H9EVhFYjWV1yNt8BS4pO8b4CH781QP8ClSB7r33',
    }
    r = requests.get(
        f"https://graph.facebook.com/v12.0/{id}",params=params)
    print(r.status_code)
    print(r.text)

# get_photo_facebook(4327063214039376)

# get_facebook_live('405370274427849')
# send_message("4368040723287885",f'{settings.MEDIA_ROOT}/image.png')

def copy(source,dest):
    db.query_commit(f'create database if not exists {dest}')
    metadata = MetaData()
    db1 = sqlalchemy.create_engine(f"mysql+pymysql://{userdb}:{passworddb}@{hostdb}/{source}")
    db2 = sqlalchemy.create_engine(f"mysql+pymysql://{userdb}:{passworddb}@{hostdb}/{dest}")
    result = db.query(f"show tables from {source}")
    result = list(result.fetchall())
    result2 = db.query(f"show tables from {dest}")
    result2 = list(result2.fetchall())
    for i in result:
        if i not in result2:
            table = sqlalchemy.Table(i[0],metadata,autoload=True, autoload_with=db1)
            table.create(db2,checkfirst=True)

# copy('muslin','Nilyn145')
def get_database():
    result = db.query('show databases')
    result = list(result.fetchall())
    return [i[0] for i in result if i[0] not in ['information_schema', 'mysql', 'performance_schema','sys','image'] ]

def get_idsell(s:str):
    try:
        first,second = s.split('-')
        if first[1].isalpha():
            first_char = first[1:]
        else:
            first_char = first[0:]
        num_char = first_char[1:]
        num_char = str(int(num_char))
        return first_char[0] + num_char
    except:
        print(s)

def get_idsell_for_qc(string:str):
    try:
        char = string.split('-')[0].split('0')[0]
        size = string.split('-')[1]
        try:
            number = int(string.split('-')[0].split(char)[1])
        except Exception as e:
            return (f'error at {string} cuz {e}')
        if len(char) > 1:
            char = char[-1]
        return char + str(number),size
    except:
        print('error at ',string)
        return string,size

def diff_day(d1, d2):
    """
        datetime(2010,10,1), datetime(2010,9,1)
    """
    return ((d1.year - d2.year) * 12 + d1.month - d2.month) * 30 + d1.day - d2.day

def write_image(imgpath,title_text):
    basewidth = 980
    my_image = Image.open(imgpath)
    if my_image.mode != 'RGB':
        my_image = my_image.convert("RGB")
    width, height = my_image.size
    width = int(width/height*basewidth)
    my_image = my_image.resize((width,basewidth), Image.ANTIALIAS)
    
    x, y = width*65/100,basewidth*10/100

    shadowcolor = "white"
    fillcolor = 'red'

    title_font = ImageFont.truetype(f'{settings.MEDIA_ROOT}/luxury_thai.ttf', 50)
    image_editable = ImageDraw.Draw(my_image)

    # image_editable.text((1000,15), title_text, (0, 0, 0),font=title_font)
    image_editable.text((x-1.5, y-1.5), title_text, font=title_font, fill=shadowcolor)
    image_editable.text((x+1.5, y-1.5), title_text, font=title_font, fill=shadowcolor)
    image_editable.text((x-1.5, y+1.5), title_text, font=title_font, fill=shadowcolor)
    image_editable.text((x+1.5, y+1.5), title_text, font=title_font, fill=shadowcolor)

    image_editable.text((x, y), title_text, font=title_font, fill=fillcolor)
    my_image.save(imgpath)

def write_image2(imgpath,imgpath2):
    my_image = Image.open(imgpath)
    width, height = my_image.size
    basewidth = int(height*1)
    width = int(width/height*basewidth)
    my_image = my_image.resize((width,basewidth), Image.ANTIALIAS)
    
    my_image.save(imgpath2)

def write_image_top_right(imgpath,title_text):
    basewidth = 800
    my_image = Image.open(imgpath)
 
    pil = ImageOps.exif_transpose(my_image)
    # ------------ crop -------------------------------

    # left = 0
    # # top = height / 5
    # # right = width
    # top = 400
    # right = 1080
    # bottom = 1480
    # my_image = my_image.crop((left, top, right, bottom))

    # ------------ crop -------------------------------
    width, height = pil.size
    # width2 = int(width/height*basewidth)

    # my_image = my_image.resize((width,basewidth), Image.ANTIALIAS)
    
    # x, y = width*75/100,basewidth*.5/100
    LEN = 85 - len(title_text)*5
    x, y = width*LEN/100,height*.5/100

    shadowcolor = "black"
    fillcolor = 'white'
    title_font = ImageFont.truetype(f'{settings.MEDIA_ROOT}/luxury2.ttf',int(width/7.6))
    image_editable = ImageDraw.Draw(pil)
    # image_editable.text((1000,15), title_text, (0, 0, 0),font=title_font)
    image_editable.text((x-int(width/7.6/40.66), y-int(width/7.6/40.66)), title_text, font=title_font, fill=shadowcolor)
    image_editable.text((x+int(width/7.6/40.66), y-int(width/7.6/40.66)), title_text, font=title_font, fill=shadowcolor)
    image_editable.text((x-int(width/7.6/40.66), y+int(width/7.6/40.66)), title_text, font=title_font, fill=shadowcolor)
    image_editable.text((x+int(width/7.6/40.66), y+int(width/7.6/40.66)), title_text, font=title_font, fill=shadowcolor)

    image_editable.text((x, y), title_text, font=title_font, fill=fillcolor)

    pil.save(imgpath,quality=100)

def crop(imgpath):
    my_image = Image.open(imgpath)
    width, height = my_image.size
    # ------------ crop -------------------------------

    left = int(width/2)
    # top = height / 5
    # right = width
    top = 0
    right = width
    bottom = int(height/5)
    my_image = my_image.crop((left, top, right, bottom))

    # ------------ crop -------------------------------
    my_image.save(imgpath,quality=100)

def write_image_top_right2(imgpath,title_text):
    my_image = Image.open(imgpath)

    if my_image.mode != 'RGB':
        my_image = my_image.convert("RGB")
    width, height = my_image.size

    x, y = width*55/100,height*.5/100


    shadowcolor = "black"
    fillcolor = 'white'

    title_font = ImageFont.truetype(f'{settings.MEDIA_ROOT}/luxury2.ttf', int(width/20))
    image_editable = ImageDraw.Draw(my_image)
    # image_editable.text((1000,15), title_text, (0, 0, 0),font=title_font)
    image_editable.text((x-1.5, y-1.5), title_text, font=title_font, fill=shadowcolor)
    image_editable.text((x+1.5, y-1.5), title_text, font=title_font, fill=shadowcolor)
    image_editable.text((x-1.5, y+1.5), title_text, font=title_font, fill=shadowcolor)
    image_editable.text((x+1.5, y+1.5), title_text, font=title_font, fill=shadowcolor)

    image_editable.text((x, y), title_text, font=title_font, fill=fillcolor)
    try:
        my_image.save(imgpath,quality=100)
    except:
        print(imgpath)

# web.update_Payment()
# web.update_fee()
def write_image_all():
    web = Web("7KRzYzjPqknzzSM2nVcooo3sWNF6EK4Oyq9QtGI8uyk=","RA9VD1AjwaHo8UW0uNk924SnxN0xIFIGdlelDEcTEE=","Muslin.info@gmail.com")
    # web.get_purchasedorder()
    imagepath = web.get("GETPRODUCTS",'imagepath','')
    sku = web.get("GETPRODUCTS",'sku','')

    sku_list = []

    for i in range(len(sku)):
        if imagepath[i]:
            if not get_idsell(sku[i]) in sku_list:
                sku_list.append(get_idsell(sku[i]))
                with open(f'image/{get_idsell(sku[i])}.jpg','wb') as f:
                    f.write(convert_url_to_bytes(imagepath[i]))
                write_image(f'image/{get_idsell(sku[i])}.jpg',get_idsell(sku[i]))

    imagepath = web.get("GETPRODUCTS",'imagepath','',2)
    sku = web.get("GETPRODUCTS",'sku','',2)

    for i in range(len(sku)):
        if imagepath[i]:
            if not get_idsell(sku[i]) in sku_list:
                sku_list.append(get_idsell(sku[i]))
                with open(f'image/{get_idsell(sku[i])}.jpg','wb') as f:
                    f.write(convert_url_to_bytes(imagepath[i]))
                write_image(f'image/{get_idsell(sku[i])}.jpg',get_idsell(sku[i]))

def concat_someColumn_by_group():
    web = Web("7KRzYzjPqknzzSM2nVcooo3sWNF6EK4Oyq9QtGI8uyk=","RA9VD1AjwaHo8UW0uNk924SnxN0xIFIGdlelDEcTEE=","Muslin.info@gmail.com")
    sku = web.get("GETPRODUCTS",'sku','')
    description = web.get("GETPRODUCTS",'description','')
    sku2 = web.get("GETPRODUCTS",'sku','',2)
    description2 = web.get("GETPRODUCTS",'description','',2)

    sku.extend(sku2)
    description.extend(description2)
    data = list(zip(sku,description))

    df = pd.DataFrame(data,columns=['SKU','description'])
    size_order = CategoricalDtype(
    ['F','XS', 'S', 'M', 'L', 'XL','2XL','3XL'], 
    ordered=True)

    df = df[df['description'] != '']
    df = df[~df["SKU"].str.contains('Q')]
    df.to_excel("V1.xlsx")
    df['Size'] = df['SKU'].apply(lambda x: x.split('-')[1])
    df['Size'] = df['Size'].astype(size_order)
    df = df.sort_values('Size')
    df['IDSell'] = df['SKU'].apply(lambda x:get_idsell(x))
    df = df[df['IDSell'].duplicated(keep=False)]
    df['DataSized'] = df.groupby('IDSell')['description'].transform(lambda x: '\n'.join(x))
    df.to_excel("V2.xlsx",header=True,index=False)

def update_sql_by_sku(sku,data,dep):
    name,breast,minwrest,maxwrest,hip,detail,data_size = data
    if not breast:
        breast = 10
    if not minwrest:
        minwrest = 10
    if not maxwrest:
        maxwrest = 10
    if not hip:
        hip = 10
    task = f"""
        update {dep}.stock_main
        set descript = '{name}', breast = '{breast}', minwrest = '{minwrest}', maxwrest = '{maxwrest}', hip = '{hip}', detail = '{detail}', data_size = '{data_size}'
        where sku = '{sku}'
    """
    db.query_commit(task)
    task = f"""
        update {dep}.data_size
        set data_size = '{data_size}'
        where sku = '{sku}'
    """
    db.query_commit(task)

def insert_sql_by_sku(data,dep):
    sku,name,breast,minwrest,maxwrest,hip,detail,data_size = data
    task = f"""
        insert into {dep}.stock_main
        values ('{sku}','{name}','{sku.split("-")[1]}','','{breast}','{minwrest}','{maxwrest}','{hip}','{detail}','{data_size}')
    """
    db.query_commit(task)
    task = f"""
        insert into {dep}.data_size
        values ('{sku}','{sku.split("-")[1]}','{get_idsell(sku)}','{data_size}')
    """
    db.query_commit(task)

def insert_data_size(sku,s,idsell,size):
    task = f"""
    insert into muslin.data_size
    values('{sku}','{s}','{idsell}','{size}')
    """
    db.query_commit(task)

def export_excel(script, name, dep='muslin'):
    cursor = db.query_custom(script, dep)
    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()
    df = pd.DataFrame(list(data), columns=columns)

    df.to_excel(f'{settings.MEDIA_ROOT}/stock/{name}.xlsx', index=False)
    return f'media/stock/{name}.xlsx'


def update_table_size():
    query = """select sku from muslin.stock_main"""
    result = db.query(query)
    result = list(result.fetchall())
    result = [i[0] for i in result]
    for i in len(result):
        size_dict = {"XS":0,"S":1,"M":2,"L":3,"XL":4,"XXL":5,"2XL":6,"3XL":7}
        task = f"""
        insert into size
        values ('{result[i]}','{result[i].split("-")[1]}','{size_dict[{result[i].split("-")[1]}]}')
        """
        db.query_commit(task)

def get_data(path,dep):
    df = pd.read_excel(path)

    if dep == 'muslin':
        column = {"sku":"SKU.1","name":"Description","amount":'Incoming ',"cost":'Unnamed: 13',"price":"Price"}
    else:
        column = {"sku":"SKU.1","name":"Description","amount":'Incoming ',"cost":'Unnamed: 10',"price":"Price"}
    df = df.iloc[1: , :]
    df.dropna(subset = [column['sku']], inplace=True)

    sku = [df.loc[i,column['sku']] for i in df.index]
    name = [df.loc[i,column['name']] for i in df.index]
    amount = [df.loc[i,column['amount']] for i in df.index]
    cost = [df.loc[i,column['cost']] for i in df.index]
    price = [df.loc[i,column['price']] for i in df.index]

    data = list(zip(sku,name,amount,cost,price))
    return data

def export_excel_vrichform(data,dest_path):
    df = pd.DataFrame(data, columns=['รหัสสินค้า', 'รายละเอียด', 'จำนวน', 'ต้นทุน','ราคา'])

    def get_cost(sku):
        try:
            sku = get_idsell(sku)[0]
            price_dict = {"A":790,"B":890,"C":950,"D":890,"E":790,"X":690,"N":790,"T":890,"P":950,"R":890,"K":890,'Y':890,"U":890,"G":890,"V":890,'J':0,'ก':690,'W':790}
            return price_dict[sku]
        except:
            return 0

    df['รหัสขาย'] = df['รหัสสินค้า'].apply(get_idsell)
    df = df[df['รหัสขาย'] != 'wrong']
    df['หน่วย'] = ''
    df['ค่าส่งเพิ่ม'] = 0
    df['หมายเหตุ'] = ''
    df['จำนวน'] = df['จำนวน'].apply(pd.to_numeric, errors='ignore')
    df['ต้นทุน'] = df['ต้นทุน'].apply(pd.to_numeric, errors='ignore')
    df['ราคา'] = df['ราคา'].apply(pd.to_numeric, errors='ignore')
    df = df[['รหัสสินค้า', 'รหัสขาย', 'รายละเอียด', 'หน่วย',
             'จำนวน', 'ราคา', 'ต้นทุน', 'ค่าส่งเพิ่ม', 'หมายเหตุ']]
    df.to_excel(f'{dest_path}', index=False)

def prepare_keyorder(dep):    
    task = "select id,min(status) from addorder_detail group by id"
    result = db.query_custom(task,dep)
    result = list(result.fetchall())
    for i in result:
        if int(i[1]) > 0:
            key(dep,i[0])

def key(dep,id):
    #product
    task = f"""select addorder_detail.*,descript,stock_main.amount from addorder_detail
            inner join stock_main on stock_main.sku = addorder_detail.sku
            where id ={id}"""
    product = db.query_custom(task,dep)
    product = list(product.fetchall())
    data_list = []
    for i in product:
        name = i[5]
        sku = i[1]
        amount = int(i[2])
        price = int(i[3])   
        web.post("UPDATEPRODUCTAVAILABLESTOCKLIST",sku,int(i[6]) + amount)
        data_form = {"sku":sku,"name":name,"number":amount,"pricepernumber":price,"discount":"0","totalprice":price * amount}
        data_list.append(data_form)
    #product

    #information
    task = f"select * from addorder where id = {id}"
    result = db.query_custom(task,dep)
    result = list(result.fetchall())
    task = f"select max(number) from deli_zort where number like '%SO%'"
    maxnumber = db.query_custom(task,dep)
    maxnumber = list(maxnumber.fetchall())[0][0]
    splitMaxnumber = maxnumber.split('-')[1]
    splitMaxnumber = int(splitMaxnumber) + 1
    maxnumber = f"SO-{splitMaxnumber}"
    idaddorder, name, addr, phone, shippingchannel, paymentmethod, discount, shippingamount, status, date, total = result
    if 'cod' in shippingchannel.lower():
        isCOD = True
        cstname += " (COD)"
        zort_form = {
            "number":maxnumber,
            'customername':cstname,
            'customerphone':phone,
            'customeraddress':addr,
            'shippingname':cstname,
            'shippingphone':phone,
            'shippingaddress':addr,
            'shippingchannel':shippingchannel,
            'isCOD':isCOD,
            "orderdate": f"{datetime.date.today()}",
            "amount": total,
            "discount": discount,
            "vattype": 3,
            "shippingamount":shippingamount ,
            "warehousecode":"W0001",
            "list":data_list
            }
    else:
        isCOD = False
        zort_form = {
            "number":maxnumber,
            'customername':cstname,
            'customerphone':phone,
            'customeraddress':addr,
            'shippingname':cstname,
            'shippingphone':phone,
            'shippingaddress':addr,
            'shippingchannel':shippingchannel,
            'isCOD':isCOD,
            "orderdate": f"{datetime.date.today()}",
            "amount": total,
            "discount": discount,
            "vattype": 3,
            "shippingamount":shippingamount ,
            "warehousecode":"W0001",
            "paymentmethod" : paymentmethod,
            "paymentamount":total,
            "list":data_list
            }
    db.query_commit(f'update {dep}.addorder set status = 1 where id = {id}')
    web = Web(get_api_register(dep,"apikey"),get_api_register(dep,"apisecret"),get_api_register(dep,"storename"))
    web.post_order(zort_form)

def cleaned_data_for_zort(data,checked_stock,dep,path=''):
    def tranprice(dep,sku):
        if dep == 'muslin':
            try:
                idsell = get_idsell(sku)[0]
                price_dict = {"J":890,"L":890,"H":990,"A":790,"B":890,"C":990,"D":890,"E":790,"X":690,"N":790,"T":890,"P":990,"R":890,"K":890,"U":890,"G":890,"V":890,'ก':690,'W':790}
                if idsell == 'Y':
                    if 'แขนสั้นขาสั้น' in sku:
                        return 790
                    elif 'แขนยาวขายาว' in sku:
                        return 950
                    else:
                        return 890
                else:
                    return price_dict[idsell]
            except:
                return 0
        else:
            return 0
    vrich = []
    
    result = db.query(f"select sku,amount from {dep}.stock")
    result = list(result.fetchall())
    cost_dict = {}
    for i in result:
        cost_dict[i[0]] = i[1]

    result = db.query(f"select sku from {dep}.reserve")
    result = list(result.fetchall())
    reserveList = [i[0].split(' ')[0] for i in result]
        
    result = db.query(f"select sku,amount,image from {dep}.stock_main")
    result = list(result.fetchall())
    amount_dict = {}
    for i in result:
        amount_dict[i[0]] = [i[1],i[2]]

    result = db.query(f"select sku,price from {dep}.cost")
    result = list(result.fetchall())
    price_dict = {}
    for i in result:
        try:
            price_dict[i[0]] = int([i[1]])
        except:
            price_dict[i[0]] = 0

    def tran(sku):
        try:
            return price_dict[sku]
        except:
            return 0
            
    def prepare_amount2(sku,value,cost,name,dep):
        if dep == 'muslin':
            if '-' in sku:
                size = sku.split('-')[1]
            else:
                size = ''
            try:
                zort_amount = int(amount_dict[sku][0])
                if amount_dict[sku][1] == 'None':
                    return 0,value
                amount = int(value) - (2-zort_amount)
                if amount >= 0:
                    return (2-zort_amount),amount
                else:
                    return int(value),0
            except:
                vrich.append(sku)
                db.query_commit(f'insert into {dep}.stock values ("{sku}",0);\n')
                db.query_commit(f'insert into {dep}.stock_detailsize values ("{sku}","","","","");\n')
                db.query_commit(f"insert into {dep}.cost values('{sku}',{cost},0,{tranprice(dep,sku)})")
                task = f'''
                insert into {dep}.data_size
                values ("{sku}","{size}","{get_idsell(sku)}","")
                  '''
                db.query_commit(task)
                command = (f'''insert into {dep}.stock_main
                                    values ("{sku}","{name}","{size}","None","34","24","46",
                                          "50","","",0)''')
                db.query_commit(command)
                # with open ("multi.txt","a",encoding="utf-8") as f:
                #     f.write(f'insert into {dep}.stock_detailsize values ("{sku}","","","","");\n')
                #     f.write(f'insert into {dep}.stock values ("{sku}",0);\n')
                return 0,int(value)
        else:
            try:
                zort_amount = int(amount_dict[sku][0])
                return int(value),0
            except:
                print('new',sku)
                if '-' in sku:
                    size = sku.split('-')[1]
                else:
                    size = ''
                # db.query_commit(f'insert into {dep}.stock values ("{sku}",0);\n')
                # db.query_commit(f'insert into {dep}.stock_detailsize values ("{sku}","","","","");\n')
                vrich.append(sku)
                db.query_commit(f'insert into {dep}.stock_detailsize values ("{sku}","","","","");\n')
                db.query_commit(f'insert into {dep}.stock values ("{sku}",0);\n')
                task = f"insert into {dep}.cost values('{sku}',{cost},0,{tranprice(dep,sku)})"
                print(task)
                db.query_commit(task)
                task = f'''
                insert into {dep}.data_size
                values ("{sku}","{size}","{get_idsell(sku)}","")
                  '''
                db.query_commit(task)
                command = (f'''insert into {dep}.stock_main
                                    values ("{sku}","{name}","{size}","None","34","24","46",
                                          "50","","",0)''')
                db.query_commit(command)

            return int(value),0

    def prepare_amount_vrich(sku):
        try:
            return cost_dict[sku]
        except:
            return 0

    data_list = []
    taskdb_stock_main = ''
    taskdb_stock = ''
    sum_all = 0
    for i in range(len(data)):
        data[i] = list(data[i])
        sku = str(data[i][0])
        if sku in reserveList:
            result = db.query_custom(f"select fname from reserve where sku like '%{sku}%'",dep)
            result = list(result.fetchall())
            for res in result:
                send_line_return(f"รหัสนี้มาแล้ว {sku} ลูกค้าชื่อ {res[0]}")
        name = str(data[i][1])
        data[i].append(tran(sku))
        if not sku in name:
            name = f"{data[i][0]} {data[i][1]}"
        
        a,b = prepare_amount2(sku,data[i][2],float(data[i][3]),str(name),dep)
        data_form = {"sku":str(sku),"name":str(name),"number":int(data[i][2]),"pricepernumber":float(data[i][3]),"discount":"0","totalprice":float(data[i][3]) * int(data[i][2])}
        data_list.append(data_form)
        sum_all += float(float(data[i][3]) * int(data[i][2]))
        if not checked_stock:
            data[i][2] += prepare_amount_vrich(data[i][0])
        if dep == 'muslin':
            try:
                taskdb_stock_main += f"WHEN '{sku}' THEN { int(amount_dict[sku][0]) + int(data[i][2])}\n"
            except:
                taskdb_stock_main += f"WHEN '{sku}' THEN 0\n"
    if dep == 'muslin':
        task_final = f'''
        UPDATE {dep}.stock_main
        SET amount
        = CASE sku
        {taskdb_stock_main}
        ELSE amount
        END;
        '''
        db.query_commit(task_final)

    if not path:
        refer = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(10))
    else:
        refer = path
    if dep == 'muslin':
        warehouse = 'W0002'
    else:
        warehouse = 'W0001'
    if checked_stock:
        warehouse = 'W0001'
        
    zort_form = {
        "reference":f"{refer} {datetime.date.today()}",
        "number": f"{refer} {datetime.date.today()}",
        "purchaseorderdate": f"{datetime.date.today()}",
        "amount": sum_all,
        "warehousecode":warehouse,
        "paymentmethod" : "Cash",
        "paymentamount":sum_all,
        "status" : "Success",
        "list":data_list
        }
    return vrich,zort_form

def get_diff_stock(path, dep):
    CHECK = False
    descript = db.query(f"select sku, descript from {dep}.stock_main")
    descript = dict(descript.fetchall())

    price_result = db.query(f"select sku, cost from {dep}.cost")
    price_dict = dict(price_result.fetchall())

    sell_result = db.query(f"select sku, price from {dep}.cost")
    sell_dict = dict(sell_result.fetchall())

    def get_cost(sku):
        return price_dict.get(sku, 0)

    df = pd.read_excel(path)

    column = [i for i in df.columns]

    sku = [df.loc[i, column[0]] for i in df.index if str(df.loc[i, column[0]]) != 'nan']

    if len(column) < 2:
        amount_dict = Counter(sku)
        sku = list(dict.fromkeys(sku))
        amount = [amount_dict[i] for i in sku]
    else:
        amount = [df.loc[i, column[1]] for i in df.index if str(df.loc[i, column[0]]) != 'nan']

    descript = [descript.get(i, i) for i in sku]
    cost, price, error_sku = [], [], []

    for i in sku:
        cost.append(get_cost(i))
        price.append(sell_dict[i])

    name = [d.replace(s, '').strip() if s in d else d for s, d in zip(sku, descript)]

    if not CHECK:
        data = list(zip(sku, name, amount, cost, price))
    else:
        print('error 2440')
        df = pd.DataFrame(error_sku, columns=['รหัสที่พิมพ์ผิด'])
        df.to_excel(f"{settings.MEDIA_ROOT}/stock/diff.xlsx", index=False)
        return f"{settings.MEDIA_ROOT}/stock/diff.xlsx"

    result = db.query_custom("select stock.sku, stock.amount + stock_main.amount from stock_main "
                             "inner join stock on stock_main.sku = stock.sku", dep)
    result_dict = dict(result.fetchall())
    data = [(sku, n, a, result_dict.get(sku, 0), result_dict.get(sku, 0) - a, c, p) for sku, n, a, c, p in data]

    skus = [i[0] for i in data]
    for sku in result_dict:
        if sku not in skus:
            data.append((sku, sku, 0, result_dict[sku], result_dict[sku], 0, 0))

    columns = ['SKU', 'ชื่อ', 'จำนวนใหม่', 'จำนวนเก่า', 'จำนวนต่าง', 'ต้นทุน', 'ราคา']
    df = pd.DataFrame(data, columns=columns)
    df['จำนวนต่าง'] = df['จำนวนต่าง'].abs()
    df = df.sort_values(by='จำนวนต่าง', ascending=False)
    df.to_excel(f"{settings.MEDIA_ROOT}/stock/diff.xlsx", index=False)

# get_diff_stock('my pajamas 19.12 final.xlsx','maruay')
def upstock(path,dep):
    def tran(dep,sku):
        if dep == 'muslin':
            try:
                idsell = get_idsell(sku)[0]
                print(f"idsell = {idsell}")
                price_dict = {"J":890,"L":890,"H":990,"A":790,"B":890,"C":990,"D":890,"E":790,"X":690,"N":790,"T":890,"P":990,"R":890,"K":890,"U":890,"G":890,"V":890,'J':'','ก':690,'W':790}
                if idsell == 'Y':
                    if 'แขนสั้นขาสั้น' in sku:
                        return 790
                    elif 'แขนยาวขายาว' in sku:
                        return 950
                    else:
                        return 890
                else:
                    return price_dict[idsell]
            except:
                return 0
        else:
            result = db.query(f"select sku,price from {dep}.cost")
            result = list(result.fetchall())
            price_dict = {}
            for i in result:
                price_dict[i[0]] = i[1]
            try:
                return price_dict[sku]
            except:
                return 0

    data = get_data(path,dep)
    # for i in data:
    # get_diff(data)
    vrich,zort = cleaned_data_for_zort(data,False,dep,os.path.split(str(path))[1])
    path = f"{str(path).split('.')[0]}.xlsx"
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    web.post_purchase_order(zort)
    data = list(data)
        
    for i in vrich:
        web.add_tag(web.getId(i),["ecom", "stock", "ภาพไม้แขวน_image", "model_image"])

    for i in data:
        print(i)
        web.updateproduct(i[0],i[4])
    
    for i in data:
        sku = i[0]
        print(f"sku : {sku}")
        tag = web.get_tag_by_sku(sku)
        if not tag:
            tag = []
        else:
            tag = list(tag)
        tag.append('สินค้ายังไม่ QC')
        web.add_tag(web.getId(sku),tag)

    return path
def check_stock(path,dep):
    descript = db.query(f"select sku,descript from {dep}.stock_main")
    descript = list(descript.fetchall())
    descript_dict = {}
    for i in descript:
        descript_dict[i[0]] = i[1]

    price = db.query(f"select sku,cost from {dep}.cost")
    price = list(price.fetchall())
    price_dict = {}
    for i in price:
        price_dict[i[0]] = i[1]

    def get_cost(sku):
        try:
            return price_dict[sku]
        except:
            return 0

    def get_data_checkstock(path):
        df = pd.read_excel(path)

        column = [i for i in df.columns]

        sku = [df.loc[i,column[0]] for i in df.index if str(df.loc[i,column[0]]) != 'nan']
        # amount = [df.loc[i,'จำนวน'] for i in df.index]

        # count duplicated
        amount_dict = Counter(sku)

        # remove duplicated
        sku = list(dict.fromkeys(sku))
 
        amount = [amount_dict[i] for i in sku]
        descript = []
        for i in sku:
            try:
                descript.append(descript_dict[i])
            except:
                descript.append(i)
        cost = [get_cost(i) for i in sku]
        data = list(zip(sku,descript,amount,cost))
        df = pd.DataFrame(data)
        df.to_excel(f"{settings.MEDIA_ROOT}/stock/checkstock_cleaned.xlsx",index=False)

        task = "select descript from RMA_data where id in (select id from RMA where recieve_time is null and order_status != 'ส่งทันที')"
        reservedForRma = db.query_custom(task,dep)
        reservedForRma = list(reservedForRma.fetchall())
        reservedForRma = [i[0] for i in reservedForRma]
        reservedSku = []
        for i in reservedForRma:
            task = db.query_custom(f"select sku from stock_main where descript = '{i}'",dep)
            try:
                sku = list(task.fetchall())[0][0]
            except:
                task = db.query_custom(f"select sku from stock_main where descript like '%{i.split(' ')[0]}%'",dep)
                sku = list(task.fetchall())[0][0]

            reservedSku.append(sku)
        for i in range(len(data)):
            if data[i][0] in reservedSku:
                data[i] = list(data[i])
                data[i][2] -= 1
                data[i] = tuple(data[i])
                print(data[i])
        print('ZZZZZZZZZZZZZZZZZZZZZZZ')
        for i in data:
            if i[0] in reservedSku:
                print(i)
        task =f"""update {dep}.addorder_detail inner join addorder on addorder.idaddorder = addorder_detail.id
            set addorder_detail.status = 0  where addorder.status = 0"""
        
        # db.query_commit(task)

        task = f"""select addorder_detail.* from addorder_detail inner join addorder on addorder.idaddorder = addorder_detail.id
                where addorder.status = 0 """
        if dep == 'muslin':
            result = db.query_custom(task,dep)
            result = list(result.fetchall())
            reserveOrder = {}
            for i in result:
                reserveOrder[i[1]] = int(i[2]),i[0]

            for i in data:
                if data[0] in reserveOrder:
                    data[2] -= reserveOrder[data[0]][0]
                    # db.query_commit(f"update {dep}.addorder_detail set status = 1 where id = {reserveOrder[data[0]][1]} and sku = '{data[0]}'")
                    print(f"update {dep}.addorder_detail set status = 1 where id = {reserveOrder[data[0]][1]} and sku = '{data[0]}'")

            # prepare_keyorder(dep)
        return data
        
    data = get_data_checkstock(path)
    print('done getdata')
    # get_diff(data,',muslin')
    vrich,zort = cleaned_data_for_zort(data,True,dep)
    print('cleaned')
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    web.post_purchase_order(zort)
    # runLoadpath()
    return "media/stock/checkstock_cleaned.xlsx"

def nospecial(text):
	import re
	text = re.sub("[^a-zA-Z0-9]+", "",text)
	return text

web = Web("7KRzYzjPqknzzSM2nVcooo3sWNF6EK4Oyq9QtGI8uyk=","RA9VD1AjwaHo8UW0uNk924SnxN0xIFIGdlelDEcTEE=","Muslin.info@gmail.com")

# web.get_minus_available('W0001')
# web.update_fee()
# web = Web("pteRXLvqBNcUXlgIB3RDHxBn3vXoi9cwRp6u/v9M=","GbJK2j7YS5dJtVaBomSsdyenjYuEwdI2A4gLPbKrRAI=","Maruay18.co.th@gmail.com")
# web.update_order_shipping()
# web.get_track_2('maruay')

def insert_jt(path,dep):
    df = pd.read_excel(path)
    colum = [i for i in df.columns]
    df = df[1:]
    df.dropna(subset = [colum[1]], inplace=True)
    
    track = [df.loc[i,colum[1]] for i in df.index]
    cstname = [df.loc[i,colum[2]] for i in df.index]
    day = datetime.datetime.now() - datetime.timedelta(days = 1)
    # day = datetime.datetime.now()
    for i in range(len(cstname)):
        db.query_commit(f"insert into {dep}.jt values ('{track[i]}','{cstname[i]}','{day.strftime('%Y-%m-%d')}')")

def get_diff_JT(day):
    task = f"select *,if(customername in (select customername from jt where shippingdate = '{day}'),'เปลี่ยนเลขแทรค','ไม่มีแทรค') as สาเหตุ from deli_zort where trackingno not in (select trackingno from jt) and shippingtime = '{day}'"
    export_excel(task,'diff2')
    task = f"select distinct(IDorder),fbname,cstname,if(cstname in (select customername from jt where shippingdate = '{day}'),'เปลี่ยนเลขแทรค','ไม่มีแทรค') as สาเหตุ from deli_vrich where trackingno not in (select trackingno from jt)"
    export_excel(task,'vrich_diff')

def post_zero_zort(dep):
    cancelLoadpath()
    db.query_commit(f"update {dep}.stock set amount = 0")
    db.query_commit(f"update {dep}.stock_main set amount = 0")
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    for i in range(3):
        try:
            task = web.post_zero(1,dep)
            task2 = web.post_zero(2,dep)
            task3 = web.post_zero(3,dep)
            task4 = web.post_zero(4,dep)
            task5 = web.post_zero(5,dep)
            task6 = web.post_zero(6,dep)
            task7 = web.post_zero(7,dep)
            task_final = f'''
            UPDATE {dep}.stock_main
            SET amount
            = CASE sku
            {task}
            {task2}
            {task3}
            {task4}
            {task5}
            {task6}
            {task7}
            ELSE amount
            END;
            '''
            db.query_commit(task_final)
        except:
            print('error')
        print('done')
# post_zero_zort('muslin')

def get_amount(sku,amount,dep='muslin'):
    amount = int(amount)
    result = db.query(f"select amount from {dep}.stock where sku = '{sku}'")
    result = list(result.fetchall())[0][0]
    result = int(result)
    new_amount = 2 - amount
    if result - new_amount >= 0:
        return 2,result - new_amount
    else:
        return amount + result,0

def upstock_no_vrich(refername,path,dep):

    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    descript = db.query(f"select sku,descript from {dep}.stock_main ")
    descript = list(descript.fetchall())
    descript_dict = {}
    for i in descript:
        descript_dict[i[0]] = i[1]

    def get_cost(sku):
        try:
            sku = get_idsell(sku)[0]
            price_dict = {"A":790,"B":890,"C":950,"D":890,"E":790,"X":690,"N":790,"T":890,"P":950,"R":890,"K":890,'Y':890,"U":890,"G":890,"V":890,'J':0,'ก':690,'W':790}
            return price_dict[sku]
        except:
            return 0

    def get_data_checkstock(path):
        df = pd.read_excel(path)
        
        columns = [i for i in df.columns]
        sku = [df.loc[i,columns[0]] for i in df.index]

        # count duplicated
        amount_dict = Counter(sku)

        # remove duplicated
        sku = list(dict.fromkeys(sku))

        amount = [amount_dict[i] for i in sku]
        descript = []
        for i in sku:
            try:
                descript.append(f"{descript_dict[str(i).strip()]}")
            except:
                descript.append(f"{i} none")
        cost = [get_cost(i) for i in sku]
        data = list(zip(sku,descript,amount,cost))
        # df.to_excel("checkstock_cleaned.xlsx",index=False)
        return data

    data = get_data_checkstock(path)

    data_list = []
    sum_all = 0
    for i in range(len(data)):
        data[i] = list(data[i])
        sku = str(data[i][0])
        name = str(data[i][1])
        amount = int(data[i][2])

        #cost
        data[i][3] = 0

        if not sku in name:
            name = f"{sku} {name}"
        
        data_form = {"sku":sku,"name":name,"number":amount,"pricepernumber":data[i][3],"discount":"0","totalprice":data[i][3] * amount}
        data_list.append(data_form)
        sum_all += data[i][3] * amount

    zort_form = {
        "reference":f"{refername}",
        "number": f"{refername}",
        "purchaseorderdate": f"{datetime.date.today()}",
        "amount": sum_all,
        "warehousecode":"W0001",
        "paymentmethod" : "Cash",
        "paymentamount":sum_all,
        "status" : "Success",
        "list":data_list
        }

    web.post_purchase_order(zort_form)

def export_to_vrich(dep,excelname):
    task = """
    select stock_main.sku,stock_main.descript,stock.amount,cost.cost,cost.price from stock_main
    inner join stock on stock_main.sku = stock.sku
    inner join cost on stock_main.sku = cost.sku
    where stock.amount > 0
    """
    result = db.query_custom(task,dep)
    result = list(result.fetchall())

    sku = [i[0] for i in result]
    descript= [i[1] for i in result]
    name = []

    for i in range(len(sku)):
        if sku[i] in descript[i]:
            name.append(descript[i].replace(sku[i],'').strip())
        else:
            name.append(descript[i])

    amount= [i[2] for i in result]
    cost = [i[3] for i in result]
    price = [i[4] for i in result]

    data = list(zip(sku,name,amount,cost,price))
    export_excel_vrichform(data,f'{excelname}.xlsx')
# export_to_vrich('muslin','ก่อนไลฟ์',True)
def manageStockLive(name):
    task = """
    SELECT a.min_descript, a.idsell, a.sum_amount, concat(a.count_sku,":", b.count_sku) as ratio
    FROM 
    (SELECT min(stock_main.descript) as min_descript, idsell , sum(stock_main.amount -4 ) as sum_amount, count(stock.sku) as count_sku
    FROM stock_main
    INNER JOIN data_size ON data_size.sku = stock_main.sku 
    INNER JOIN stock ON stock.sku = stock_main.sku 
    WHERE stock_main.amount > 4
    GROUP BY idsell) as a
    LEFT JOIN (SELECT idsell, count(sku) as count_sku from data_size group by idsell) as b 
    ON a.idsell = b.idsell
    order by idsell ,ratio
    """
    export_excel(task,name)
# manageStockLive('live')
def send_line_blog(msg):
    url = 'https://notify-api.line.me/api/notify'
    token = 'b3pgaYmoqCp51kMPzL4pgknKg68Gc4227OxVg3y3nVe'
    headers = {'content-type': 'application/x-www-form-urlencoded',
               'Authorization': 'Bearer '+token}
    r = requests.post(url, headers=headers, data={'message': msg})

def send_line_blog_admin(msg):
    url = 'https://notify-api.line.me/api/notify'
    token = 'TcfnJ46PQkvm2AqOHhpurxvOCVczm8ruZ023EYjSBb6'
    headers = {'content-type': 'application/x-www-form-urlencoded',
               'Authorization': 'Bearer '+token}
    r = requests.post(url, headers=headers, data={'message': msg})
 
def table_break(time:str,name):
    task = """
    SELECT 
    name,
    DATE(checkin),
    TIME(DATE_ADD(checkin, INTERVAL 7 HOUR)) AS 'เวลาเข้างาน',
    TIME(DATE_ADD(checkout, INTERVAL 7 HOUR)) AS 'เวลาออกงาน',
    TIME(DATE_ADD(breakin, INTERVAL 7 HOUR)) AS 'เวลาเข้าเบรค',
    TIME(DATE_ADD(breakout, INTERVAL 7 HOUR)) AS 'เวลาออกเบรค',
    IF(HOUR(TIMEDIFF(checkin, checkout)) - 1 >= 8, 8, HOUR(TIMEDIFF(checkin, checkout)) - 1) AS 'ชั่วโมงทำงาน',
    IF(TIMEDIFF(breakin, breakout) < -010000, 'เบรคเกิน', 'เบรคพอดี') AS เช็คเบรค,
    IF(TIME(checkin) > '09:15:00', 'สาย', '') AS เช็คมาสาย
    FROM 
        face_check
    WHERE 
        checkin > DATE_SUB(NOW(), INTERVAL 2 MONTH)
    ORDER BY 
        name, checkin;
    """
    export_excel(task,name)

def reportrma(name,dep):
    task = f"""
    select max(RMA.id) as id,max(RMA.number) as ชื่อรายการ,max(reason) as สาเหตุ,max(platform) as platform,max(key_time) as เวลาที่คีย์,max(recieve_time) as เวลารับสินค้า ,group_concat(distinctrow RMA_before.descript) as สินค้าเก่า,group_concat(distinctrow RMA_data.descript) as สินคเาใหม่ from RMA 
    join RMA_before on RMA.id = RMA_before.id
    join RMA_data on RMA.id = RMA_data.id
    group by RMA.id;
    """
    export_excel(task,name,dep)
    
def getBalanceForAccount(text):
    task = f"""
    select deli_zort.idorder,max(number),max(orderdate) as วันที่,sum(order_main.amount) as จำนวนสินค้า,sum(order_main.price) as 'ราคาต่อออเดอร์',sum(T.cost) as 'ต้นทุนรวมต่อออเดอร์',sum(order_main.price) - sum(cost) as 'กำไร' from deli_zort
    inner join order_main on deli_zort.idorder = order_main.idorder
    inner join muslin.T on muslin.T.sku = order_main.sku
    where orderdate > '2022-04-01' and orderdate <'2022-06-01' and number like 'SO-%' and status != "Voided" and number not like '%return%'
    group by deli_zort.idorder order by max(orderdate) ;
    """
    export_excel(task,'zortBalance')
    task = f"""
    select idorder,max(date) as วันที่,sum(amount) as จำนวนสินค้า,sum(deli_vrich.price) as 'ราคาต่อออเดอร์',sum(T.cost) as 'ต้นทุนรวมต่อออเดอร์',sum(deli_vrich.price) - sum(cost) as 'กำไร' from deli_vrich inner join T on T.sku = deli_vrich.sku
    where date > '{text}'
    group by idorder order by max(date) ;
    """
    export_excel(task,'vrichBalance')
    task ="""
    select date(orderdate) as วันที่ ,sum(order_main.amount) as จำนวนสินค้า,count(distinct(order_main.idorder)) as จำนวนออเดอร์,sum(order_main.price) - sum(cost) as กำไรต่อวัน from deli_zort 
    inner join order_main on order_main.idorder = deli_zort.idorder
    inner join T on T.sku = order_main.sku
    where status != "Voided" and order_main.price - cost  > 0
    group by date(orderdate);
    """
    export_excel(task,'OrderZortGroupByDate')
    task = """
    select date(date) as วันที่ ,sum(deli_vrich.amount) as จำนวนสินค้า,count(distinct(IDorder)) as จำนวนออเดอร์,sum(deli_vrich.price) - sum(cost) as กำไรต่อวัน from deli_vrich
    inner join T on T.sku = deli_vrich.sku
    group by date(date);
    """
    export_excel(task,'OrderVrichGroupByDate')

def fullFill(path,destname,dep):
    CHECK = False
    descript = db.query(f"select sku,descript from {dep}.stock_main")
    descript = list(descript.fetchall())
    descript_dict = {}
    for i in descript:
        descript_dict[i[0]] = i[1]

    price = db.query(f"select sku,cost from {dep}.cost")
    price = list(price.fetchall())
    price_dict = {}
    for i in price:
        price_dict[i[0]] = i[1]

    sell = db.query(f"select sku,price from {dep}.cost")
    sell = list(sell.fetchall())
    sell_dict = {}
    for i in sell:
        sell_dict[i[0]] = i[1]

    def get_cost(sku):
        try:
            return price_dict[sku]
        except:
            return 0
            
    df = pd.read_excel(path)

    column = [i for i in df.columns]

    sku = [df.loc[i,column[0]] for i in df.index if str(df.loc[i,column[0]]) != 'nan']
    # amount = [df.loc[i,'จำนวน'] for i in df.index]

    # count duplicated
    if len(column) < 2:
        amount_dict = Counter(sku)

        # remove duplicated
        sku = list(dict.fromkeys(sku))

        amount = [amount_dict[i] for i in sku]
    else:
        amount = [df.loc[i,column[1]] for i in df.index if str(df.loc[i,column[0]]) != 'nan']
        
    descript = []
    for i in sku:
        try:
            descript.append(descript_dict[i])
        except:
            descript.append(i)
    cost,price,error_sku = [],[],[]
    for i in sku:
        try:
            cost.append(get_cost(i))
            price.append(sell_dict[i])
        except:
            CHECK = True
            error_sku.append(i)

    name = []
    for i in range(len(sku)):
        if sku[i] in descript[i]:
            name.append(descript[i].replace(sku[i],'').strip())
        else:
            name.append(descript[i])

    if not CHECK:
        data = list(zip(sku,name,amount,cost,price))
        export_excel_vrichform(data,f'{settings.MEDIA_ROOT}/stock/{destname}.xlsx')
        return f'media/stock/{destname}.xlsx'
    df = pd.DataFrame(error_sku,columns=['รหัสที่พิมพ์ผิด'])
    df.to_excel(f"{settings.MEDIA_ROOT}/stock/{destname}.xlsx",index=False)
    return f'media/stock/{destname}.xlsx'

# check_stock('เช็คสต็อกร้านเล็ก 17-11-22.xlsx','maruay')
def tran_to_vrich(excelpath):
    dep = 'muslin'
    df = pd.read_excel(excelpath)
    sku = tuple([df.loc[i,'sku'] for i in df.index])
    idsell_tup = tuple([df.loc[i,'sku'] for i in df.index])
    amount_list = []
    datadict = {}
    datadict['stocks'] = list()
    taskdb_stock = ''
    taskdb_stock_main = ''
    SKU = []
    task = f"select stock.sku,stock.amount + stock_main.amount from stock\
            inner join stock_main on stock.sku = stock_main.sku \
        inner join data_size on data_size.sku = stock.sku\
            where data_size.idsell in {idsell_tup}"
    result = db.query_custom(task,'muslin')
    amount = list(result.fetchall())
    for index in amount:
        print(index[0],index[1])
        datadict['stocks'].append({"sku": index[0], "stock": 0,'cost':0})
        taskdb_stock += f"WHEN '{index[0]}' THEN 0\n"
        taskdb_stock_main += f"WHEN '{index[0]}' THEN 0\n"
        # db.query_commit(f"update muslin.stock set amount = 0 where sku = '{index[0]}'")
        # db.query_commit(f"update muslin.stock_main set amount = 0 where sku = '{index[0]}'")
        SKU.append(index[0])
        # if index[0] in sku:
        #     amount_list.append(index[1] + 1)
        amount_list.append(index[1])
        if len(datadict['stocks']) % 300 and len(datadict['stocks']) >= 300:
            web.postzero(datadict)
            datadict['stocks'] = list()
    
    web.postzero(datadict)
    datadict['stocks'] = list()

    task_final = f'''
        UPDATE {dep}.stock_main
        SET amount
        = CASE sku
        {taskdb_stock_main}
        ELSE amount
        END;
        '''
    db.query_commit(task_final)
    task_final = f'''
        UPDATE {dep}.stock
        SET amount
        = CASE sku
        {taskdb_stock}
        ELSE amount
        END;
        '''
    db.query_commit(task_final)

    df2 = pd.DataFrame(amount)
    df2.to_excel('ตัวโชว์ 03082023 final2.xlsx',index=False)

# tran_to_vrich('ตัวโชว์ 03082023.xlsx')
# path = 'slip'
# for i in os.listdir(path):
#     print("{:.2f} %".format(os.listdir(path).index(i) / (len(os.listdir(path))-1)*100))
#     write_image_top_right2(os.path.join(path,i),i.split('.')[0])

def test2():
    task = 'select stock_main.sku,stock_main.amount  from stock_main where amount = -1'
    data = db.query_custom(task,'muslin')
    data = data.fetchall()
    taskdb_stock = ''
    taskdb_stock_main = ''
    datadict = {}
    datadict['stocks'] = list()
    for i in data:
        taskdb_stock_main += f"WHEN '{i[0]}' THEN 0\n"
        datadict['stocks'].append({"sku": i[0], "stock": 1,'cost':0})
        # db.query_commit(f"update {dep}.stock set amount = {data[i][2]} where sku = '{sku}';\n")
        # web.post("UPDATEPRODUCTAVAILABLESTOCKLIST",i[0],0 )
        # db.query_commit(f"update {dep}.stock_main set amount = 0 where sku = '{i[0]}';\n")
    
    task_final_main = f'''
    UPDATE muslin.stock_main
    SET amount
    = CASE sku
    {taskdb_stock_main}
    ELSE amount
    END;
    '''
    db.query_commit(task_final_main)
    web.postzero(datadict)
# test2() 

# UPDATE AMOUNT FROM VRICH TO ZORT AND UPDATE AMOUNT FROM VRICH TO 0
def update_vrich(dep):
    task = f'select sku,amount from {dep}.stock_vrich where amount > 0'
    result = list(db.query_custom(task,dep).fetchall())
    taskdb_stock = ''
    for i in result:
        taskdb_stock += f"WHEN '{i[0]}' THEN amount + {i[1]}\n"   
                # db.query_commit(f"update {dep}.stock set amount = {data[i][2]} where sku = '{sku}';\n")

    task_final = f'''
    UPDATE {dep}.stock
    SET amount
    = CASE sku
    {taskdb_stock}
    ELSE amount
    END;
    '''

    db.query_commit(task_final)
    task = f'update {dep}.stock_vrich set amount = 0'
    db.query_commit(task)
# update_vrich()

# GET WHICH SKU THAT SIZE SKIP
def get_no_size(dep='muslin'):
    task = """
    select data_size.idsell from data_size inner join stock_main on stock_main.sku = data_size.sku
    inner join stock on stock.sku = stock_main.sku
        group by data_size.idsell having sum(stock_main.amount + stock.amount) > 0"""
    idsell = db.query_custom(task,dep)
    idsell = list(idsell.fetchall())
    idsell = [i[0] for i in idsell]
    task = f"""
    select stock_main.sku,breast,hip from stock_main 
    inner join data_size on data_size.sku = stock_main.sku
    where data_size.idsell in {tuple(idsell)}
    order by data_size.idsell,FIELD(stock_main.size,  'F','XXS', 'XS', 'S', 'M', 'L', 'XL', '2XL','3XL', '4XL','5XL','6XL')
    """
    result = db.query_custom(task,dep)

    result = list(result.fetchall())

    for i in range(len(result)):
        if i == 0:
            continue
        if get_idsell(result[i][0]) == get_idsell(result[i - 1][0]):
            try:
                if abs(int(result[i][2]) - int(result[i-1][2])) > 2:
                    print(int(result[i][2]),int(result[i - 1][2]))
                    print(result[i][0],result[i - 1][0])
            except:
                pass
# get_no_size()
def QC(dep,path):
    df = pd.read_excel(path)
    # idsell = db.query_custom(idsell,'muslin')
    # idsell = list(idsell.fetchall())
    # for i in idsell:
    #     web.post("UPDATEPRODUCTAVAILABLESTOCKLIST",i[0],0 )
        # with open('multi.txt','a') as f:
        #     f.write(f"update muslin.stock set amount = 0 where sku = '{i[0]}';\n")
        #     f.write(f"update muslin.stock_main set amount = 0 where sku = '{i[0]}';\n")
        # sku.append(i[0])
        # amount.append(i[1])
    
    # df = pd.DataFrame(list(zip(sku,amount)))
    # df.to_excel('test.xlsx',index=False)
    old_SKU = tuple([df.loc[i,'รหัสสินค้า'] for i in df.index])
    # sku = tuple([df.loc[i,'Q_sku'] for i in df.index])
    idsell = [df.loc[i,'รหัสขาย'] for i in df.index]
    data_dict ={}
    result = db.query_custom(f"select sku,data_size from data_size where sku in {old_SKU}",dep)
    db.query_commit(f'delete from muslin.data_size where sku in {old_SKU}')
    db.query_commit(f'delete from muslin.cost where sku in {old_SKU}')
    result = list(result.fetchall())
    for i in result:
        print(i)
        data_dict[i[0]] = i[1]
    for i in range(len(old_SKU)):
        with open ("multi.txt","a",encoding="utf-8") as f:
            f.write(f"""insert into muslin.data_size values ("{old_SKU[i]}","{old_SKU[i].split("-")[1]}","M{get_idsell(old_SKU[i])}",'{data_dict[old_SKU[i]]}');\n""")
            f.write(f"""insert into muslin.cost values ("{old_SKU[i]}",0,0,390);\n""")
# path = 'stock.xlsx'
# QC('maruay',path)

def sql(empdata, table, databasedb):
    start_time = time.time()
    engine = create_engine("mysql+pymysql://" + userdb + ":" + passworddb + "@" + hostdb + "/" + databasedb)
    empdata.to_sql(table, engine, index=False, if_exists='replace', method='multi')
    # print("--- %s seconds ---" % (time.time() - start_time))

def Start(path, database):
    df = pd.read_excel(path)
    df['size'] = df.iloc[:, 0].astype(str).str.split('-', expand=True)[1]
    df = df.set_axis(['sku', 'idsell', 'descript', 'unit', 'amount', 'price', 'cost', 'fee', 'etc', 'position', 'size'], axis=1)
    df.drop(['unit', 'price', 'cost', 'fee', 'etc', 'position'], axis='columns', inplace=True)
    sql(df, 'stock_vrich', database)

# 
def download_image_vrich():
    task = 'select sku,image from stock_main where image != "none"'
    result = db.query_custom(task,'muslin')
    result = list(result.fetchall())
    sku_list = []
    for i in result:
        if i[0].split('-')[0] not in sku_list:
            with open(rf'C:\Users\Chino\Desktop\rose\{i[0].split("-")[0]}.png', 'wb') as f:
                f.write(convert_url_to_bytes(i[1]))
                f.close()
            sku_list.append(i[0].split('-')[0])
            print("{:.2f} %".format(result.index(i) / len(result) * 100))

def StartTest(path,database):
    print(path)
    df = pd.read_excel(path)
    df = df.set_axis(['IDorder','a','date', 'FBName', 'cstname', 'addr','b','c', 'tel','d','e', 'trackingNo','f','g', 'total_amount', 'cash', 'discount', 'deli_fee', 'total', 'Ebank', 'paid', 'timepaid', 'h','thing1', 'thing2', 'idsell', 'descript', 'Comment', 'amount', 'price', 'timedate', 'Printed', 'Checkout', 'Checkout_Time', 'sku'], axis=1)
    df.drop(['a','b','c','d','e','f','g','h'],axis='columns',inplace=True)
    sql(df,'deli_vrich',database)
    db.query_commit('update muslin.deli_vrich set trackingno = "" where trackingno is null ;')
#
def fullFillKorkai(path,destname,dep):
    descript = db.query(f"select sku,descript from {dep}.stock_main")
    descript = list(descript.fetchall())
    descript_dict = {}
    for i in descript:
        descript_dict[i[0]] = i[1]

    price = db.query(f"select sku,cost from {dep}.cost")
    price = list(price.fetchall())
    price_dict = {}
    for i in price:
        price_dict[i[0]] = i[1]

    sell = db.query(f"select sku,390 from {dep}.cost")
    sell = list(sell.fetchall())
    sell_dict = {}
    for i in sell:
        sell_dict[i[0]] = i[1]

    def get_cost(sku):
        try:
            return price_dict[sku]
        except:
            return 0
            
    df = pd.read_excel(path)

    column = [i for i in df.columns]

    sku = [df.loc[i,column[0]] for i in df.index if str(df.loc[i,column[0]]) != 'nan']
    # amount = [df.loc[i,'จำนวน'] for i in df.index]

    # count duplicated
    if len(column) < 2:
        amount_dict = Counter(sku)

        # remove duplicated
        sku = list(dict.fromkeys(sku))

        amount = [amount_dict[i] for i in sku]
    else:
        amount = [df.loc[i,column[1]] for i in df.index if str(df.loc[i,column[0]]) != 'nan']
        
    descript = []
    for i in sku:
        try:
            descript.append(descript_dict[i])
        except:
            descript.append(i)
    cost = [get_cost(i) for i in sku]
    price = [390 for i in range(len(sku))]

    name = []

    for i in range(len(sku)):
        if sku[i] in descript[i]:
            name.append(descript[i].replace(sku[i],'').strip())
        else:
            name.append(descript[i])

    data = list(zip(sku,name,amount,cost,price))
    export_excel_vrichform(data,f'{destname}')
    return f'{settings.MEDIA_ROOT}/stock/{destname}.xlsx'

def formKorKai(excelpath):
    df = pd.read_excel(excelpath)
    skus = [df.loc[i,'รหัสสินค้า'] for i in df.index]
    descripts = [df.loc[i,'รายละเอียด'] for i in df.index]
    amounts = [df.loc[i,'จำนวน'] for i in df.index]
     
    data_list = []

    for i in range(len(skus)):
        sku = skus[i]
        name = descripts[i]
        amount = int(amounts[i])
        price = 0

        data_form = {"sku":sku,"name":name,"number":amount,"pricepernumber":price,"discount":"0","totalprice":price * amount}
        data_list.append(data_form)

    zort_form = {
    "number":f"VR ก. วันที่ {datetime.date.today()} 2",
    'customername':"VR",
    "orderdate": f"{datetime.date.today()}",
    "amount": 0,
    "warehousecode":"W0001",
    "list":data_list
    }
    web = Web("pteRXLvqBNcUXlgIB3RDHxBn3vXoi9cwRp6u/v9M=","GbJK2j7YS5dJtVaBomSsdyenjYuEwdI2A4gLPbKrRAI=","Maruay18.co.th@gmail.com")
    web.post_order(zort_form)

def Korkai(filename):
    today = datetime.date.today()
    print(filename)
    fullFillKorkai(filename,filename,'maruay')
    df = pd.read_excel(filename)

    data_size = db.query(f"select sku,data_size from maruay.stock_main")
    data_size = list(data_size.fetchall())
    data_size_dict = {}
    for i in data_size:
        data_size_dict[i[0]] = i[1]

    idsell = []
    cost_task = []
    insert_task = []
    for i in df.index:
        sku = df.loc[i,'รหัสสินค้า']
        cost_task.append((sku,0,0,390))
        if  df.loc[i,'รหัสขาย'] not in idsell:
            idsell.append(df.loc[i,'รหัสขาย'])
            df.loc[i,'รหัสขาย'] = "ก" + str(len(idsell))
            insert_task.append((sku,sku.split('-')[1],"ก" + str(len(idsell)),data_size_dict[sku]))
        else:
            df.loc[i,'รหัสขาย'] = "ก" + str(idsell.index(df.loc[i,'รหัสขาย']) + 1)
            insert_task.append((sku,sku.split('-')[1],"ก" + str(len(idsell)),data_size_dict[sku]))
    df.to_excel(f'{settings.MEDIA_ROOT}/stock/K {today}.xlsx',index=False)
    formKorKai(f'{settings.MEDIA_ROOT}/stock/K {today}.xlsx')
    insert_task = ', '.join([str(i) for i in insert_task])
    cost_task = ', '.join([str(i) for i in cost_task])
    db.query_commit(f'delete from muslin.data_size where idsell like "%ก%"')
    task = f'insert into muslin.data_size values {insert_task}'
    db.query_commit(task)

    db.query_commit(" delete from muslin.cost where sku in (select sku from muslin.data_size where idsell like '%ก%');")
    task = f'insert into muslin.cost values {cost_task}'
    db.query_commit(task)
    
    return f'media/stock/K {today}.xlsx'

def convertQcToNormal(path,dep):
    df = pd.read_excel(path)
    realsku =[]
    for i in df.index:
        sku = df.loc[i,'sku']
        task = f'select sku from stock_main where sku like "%{sku[1:]}"'
        result = db.query_custom(task,dep)
        result = list(result.fetchall())
        try:
            realsku.append(result[0][0])    
        except:print('error',sku)
        if len(result) != 1:print('เจอหลาย ',sku)

    df = pd.DataFrame(realsku)
    df.to_excel('realsku.xlsx',index=False)

#
def cancel_COD_monthly(excelpath):
    vrich = pd.read_excel(excelpath,sheet_name='EMS Vrich')
    column = [i for i in vrich.columns]
    for i in vrich.index:
        if str(vrich.loc[i,column[-1]]) == 'cancel COD':
            print(vrich.loc[i,column[0]])

    zortmuslin = pd.read_excel(excelpath,sheet_name='EMS Muslin Zort')
    column = [i for i in zortmuslin.columns]
    for i in zortmuslin.index:
        if str(zortmuslin.loc[i,column[-1]]) == 'cancel COD':
            print(zortmuslin.loc[i,column[0]])

    zortmaruay = pd.read_excel(excelpath,sheet_name='EMS Maruay Zort')
    column = [i for i in zortmaruay.columns]
    for i in zortmaruay.index:
        if str(zortmaruay.loc[i,column[-1]]) == 'cancel COD':
            print(zortmaruay.loc[i,column[0]])

# cancel_COD_monthly('monthly sales report Feb  17.3.2023.xlsx')
# graph = facebook.GraphAPI(access_token)
access_tok = 'EAADs69tGsMEBO5GenCnyYZCQTLZAFpnmfow9KZAihhtmQxqZAb0tNmM2BgJ91UYYK2IckvrJrZAnsxLFqT7eo3hA9I6lLhZCXcLCYhaZAL6wozru5ULYgDxOUlIZA6slofrZAzVqTZBtMslZAnYRZC6vjserTKgOWLATo9ZBtuDgF2GTvZCzNZBB8qW9BhPvAZDZD'

# muslin ads account list
ad_acc_list = [
"1233977980713272",
"3126688860975489",
"736282584293594",
"3703607256535041",
"1789954911466934",
"1056201798412529",
"803785237485233",
"2703565199777489",
]

# maruay ads account list
ad_acc_list_maruay = [
"521471879981229",
"1025682725525082",
"2943245342634568",
]
# get spend from add id
def get_ads_spend(access_token, ad_acc_id, date_preset='yesterday'):
    url = f'https://graph.facebook.com/v16.0/act_{ad_acc_id}/insights'
    params = {
        'access_token': access_token,
        'fields': 'spend,account_id,account_currency,account_name,ad_id,ad_name,adset_id,adset_name',
        'date_preset': date_preset
    }

    response = requests.get(url, params=params)
    data = response.json()
    
    try:
        spend = data['data'][0]['spend']
        account_name = data['data'][0]['account_name']
        return round(float(spend), 2)
    except:
        return 0

from linebot import LineBotApi
from linebot.models import TextSendMessage
def auto_send_8am():
    channel_access_token_daily_report = 'w6F1ffyyanDJ+PMtmekbkLiKyNqQID1cWIM1u9oKwdRymskGI9BMCEplfSDsueuv/zOwv401JLWIAYNXucK6E3CuGnZWwTJxMgi91cIaY9L0tVYMPcdW3VuYDr3eEgJ+p6/bzcIeNf+21naBySayUwdB04t89/1O/w1cDnyilFU='

    spend_list = [get_ads_spend(access_tok, acc_id,'yesterday') for acc_id in ad_acc_list]
    ads_amount = round(sum(spend_list),2)
    spend_list = [get_ads_spend(access_tok, acc_id,'yesterday') for acc_id in ad_acc_list_maruay]
    ads_amount_jj = round(sum(spend_list),2)
    
    dep = 'muslin'
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    sales_amount = web.send_sales_report(dep)
    dep = 'maruay'
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    sales_amount_jj = web.send_sales_report(dep)
    print(f"ads_amount = {ads_amount} type : {type(ads_amount)}")
    print(f"sales_amount = {sales_amount} type : {type(sales_amount)}")
    percentage = round((ads_amount / sales_amount) * 100,2)
    percentage_jj = round((ads_amount_jj / sales_amount_jj) * 100,2)

    date = (datetime.datetime.now() - datetime.timedelta(days = 1)).strftime('%Y-%m-%d')
    text = f"""Sales Report {date}\nmuslin Ads : {format(ads_amount,',')}\nmuslin sales : {format(sales_amount,',')}\njj Ads : {format(ads_amount_jj,',')}\njj sales : {format(sales_amount_jj,',')}\nmuslin percent Ads : {percentage}%\njj percent Ads : {percentage_jj}%"""
    line_bot_api_daily_report = LineBotApi(channel_access_token_daily_report)
    message = TextSendMessage(text=text)

    # line_bot_api_daily_report.push_message('C7f9c7403440cef77ffb4561a74b58013',message)
    print(text)
    
def get_text_after_alphabet(text):
    match = re.search(r'[a-zA-Z]+(.*)', text)
    if match:
        return match.group(0)
    else:
        return text

import requests
import matplotlib.pyplot as plt
import io
import base64
import numpy as np

def upload_image_to_imgbb(image_stream, filename, api_key):
    url = 'https://api.imgbb.com/1/upload'
    
    # Encode the image data to base64
    image_base64 = base64.b64encode(image_stream.getvalue()).decode('utf-8')
    
    payload = {
        'key': api_key,
        'image': image_base64
    }
    
    response = requests.post(url, payload)
    
    if response.status_code == 200:
        data = response.json().get('data')
        return data.get('url')  # Return the URL of the uploaded image
    else:
        print('Failed to upload the image. Error:', response.text)
        return None

def send_sales_report(channel1_amount, channel2_amount, channel3_amount):
    # Generate sales data for the donut chart
    channels = ['Channel 1', 'Channel 2', 'Channel 3']
    sales_data = [channel1_amount, channel2_amount, channel3_amount]

    # Calculate total sales amount
    total_sales = sum(sales_data)

    sales_percentages = [sales_amt / total_sales * 100 for sales_amt in sales_data]
    sales_amounts = [f'{channels[i]}\n{sales_amt} ({sales_percentage:.1f}%)' for i, (sales_amt, sales_percentage) in enumerate(zip(sales_data, sales_percentages))]

    # Create a donut chart
    fig, ax = plt.subplots()
    ax.pie(sales_data, labels=sales_amounts, autopct='%1.1f%%', startangle=90)
    ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle
    plt.title('Sales Distribution')
    
    # Save the donut chart to a byte stream
    image_stream = io.BytesIO()
    plt.savefig(image_stream, format='png')
    image_stream.seek(0)  # Move the stream's cursor to the beginning
    
    # Upload the image to ImgBB
    imgbb_api_key = '617a5159bb15cb66953a95b1e465b76d'
    sales_report_url = upload_image_to_imgbb(image_stream, 'sales_report.png', imgbb_api_key)
    
    if sales_report_url:
        # Prepare the Line Bot request
        url = 'https://api.line.me/v2/bot/message/push'
        headers = {
            'Authorization': 'Bearer YOUR_CHANNEL_ACCESS_TOKEN',
            'Content-Type': 'application/json'
        }
        payload = {
            'to': 'USER_ID',  # Replace with the Line user ID you want to send the message to
            'messages': [
                {
                    'type': 'image',
                    'originalContentUrl': sales_report_url,
                    'previewImageUrl': sales_report_url
                },
                # ... (same for the donut chart)
            ]
        }

        # Send the Line Bot request
        response = requests.post(url, headers=headers, json=payload)

        # Close the figure to free up memory
        plt.close(fig)

        if response.status_code == 200:
            print('Sales report and donut chart sent successfully!')
        else:
            print('Failed to send the sales report and donut chart. Error:', response.text)

# send_sales_report(10000,2000,3000)
def edit_po(dep):
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    df = pd.read_excel(rf'D:\Dropbox\MUSLIN PAJAMAS ALL IN FO\Stock\ของเข้าร้านคนแก่\เดือนกรกฎาคม\23.07.2023\H2306118226 po.xlsm')
    for i in df.index:
        sku = df.loc[i,'SKU.1']
        name = f"{sku} {df.loc[i,'Description']}"
        if str(sku) != 'nan':
            db.query_commit(f'update {dep}.stock_main set descript = "{name}" where sku = "{sku}"')
            web.updateproductName(sku,name)

# edit_po('Nilyn145')

def send_databasename():
    res = list(db.query("select department from store_api").fetchall())
    return [i[0] for i in res]

def replace_and_split(text):
    # Define regular expression pattern to match the required format
    pattern = r'Size\s+(\S+)\s*รอบอก\s+(\d+["”])\s*เอว\s+(\d+-\d+["”])\s*สะโพก\s+(\d+["”])\s*เสื้อยาว\s+(\d+\.?\d*["”]?)\s*กางเกงยาว\s+(\d+\.?\d*["”]?)\s*เป้ายาว\s+(\d+\.?\d*["”]?)\s*รอบขา\s+(\d+\.?\d*["”]?)'

    # Find all matches in the text using regular expression
    matches = re.findall(pattern, text)

    # Create a new list with modified strings
    result = []
    for match in matches:
        size, chest, waist, hip, shirt_length, pants_length, inseam, leg_circumference = match
        new_text = f"{size}อก {chest}เอว {waist}สพ {hip}เกงยาว {pants_length}"
        result.append(new_text)

    return '\n'.join(result)

def tran_price(dep,sku):
    if dep == 'muslin':
        try:
            idsell = get_idsell(sku)[0]
            price_dict = {"J":890,"L":890,"H":990,"A":790,"B":890,"C":990,"D":890,"E":790,"X":690,"N":790,"T":890,"P":990,"R":890,"K":890,"U":890,"G":890,"V":890,'J':'','ก':690,'W':790}
            if idsell == 'Y':
                if 'แขนสั้นขาสั้น' in sku:
                    return 790
                elif 'แขนยาวขายาว' in sku:
                    return 950
                else:
                    return 890
            else:
                return price_dict[idsell]
        except:
            print(f'error idsell = {sku}')
            return 0
    else:
        result = db.query(f"select sku,price from {dep}.cost")
        result = list(result.fetchall())
        price_dict = {}
        for i in result:
            price_dict[i[0]] = i[1]
        try:
            return price_dict[sku]
        except:
            return 0

def checkstock_zort_all(path,dep):
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    df = pd.read_excel(path)
    datadict = {}
    datadict['stocks'] = list()
    taskdb_stock = ''

    for i in df.index:
        sku = df.loc[i,'sku']
        amount = df.loc[i,'amount']

        if i % 200 == 0:
            web.postzero(datadict)
            datadict['stocks'] = list()
        datadict['stocks'].append({"sku": sku, "stock": int(amount),'cost':0})
            # self.post("UPDATEPRODUCTSTOCKLIST",data['list'][i]['sku'],0)
        taskdb_stock += f"WHEN '{sku}' THEN {int(amount)}\n"   

    web.postzero(datadict)

    task_final = f'''
    UPDATE {dep}.stock_main
    SET amount
    = CASE sku
    {taskdb_stock}
    ELSE amount
    END;
    '''
    with open('checkstock 08082023.txt','w') as f:
        f.write(task_final)
    db.query_commit(task_final)

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
import os

def get_highlight(directory_path):
    def map_color_index_to_rgb(index):
        if index < len(COLOR_INDEX):
            return COLOR_INDEX[index]
        return None

    def process_directory(directory_path, sheet_name):
        for root, _, files in os.walk(directory_path):
            for filename in files:
                if filename.endswith(".xlsm"):
                    file_path = os.path.join(root, filename)
                    wb = load_workbook(file_path, data_only=True)
                    sh = wb[sheet_name]
                    
                    index = 3  # Start from row 3 (A3)
                    while index < 120:
                        color_index = sh[f"A{index}"].fill.start_color.index
                        if color_index == 'FFFFFF00':
                            highlighted_value = sh[f"C{index}"].value
                            text = f"value in {file_path.replace(directory_path, '').replace(os.path.sep, '/')} - Row {index}, Column C: {highlighted_value}\n"
                            with open('po_highlight.txt', 'a+') as f:
                                f.write(text)
                        index += 1

    # Set the directory path and sheet name
    sheet_name = "from"

    # Process the directory and extract highlighted values
    process_directory(directory_path, sheet_name)
    # -----------------------------------------------------------------------------------------
    # make a excel file 
    idsell = [
    "SK1210",
    "SN1210",
    "ST1210",
    "SN1211",
    "ST1211",
    "SK1211",
    "SN1212",
    "ST1212",
    "SK1212",
    "ST1213",
    "SN1213",
    "SN1214",
    "ST1214",
    "SN1215",
    "ST1215",
    "SP1215",
    "SN1216",
    "ST1216",
    "SP1216",
    "Y0109",
    "Y0110",
    "Y0111",
    "Y0112",
    "SV1219",
    "SK1219",
    "SW1219",
    "SV1217",
    "SL1217",
    "SW1217",
    "ST1218",
    "SN1218",
    "SK1218",
    "SU1098",
    "SW1220",
    "ST1221",
    "SN1221",
    "SV1220",
    "SL1220",
    "SL1221",
    "SP1183",
    "SW1222",
    "SV1222",
    "ST1223",
    "SN1223",
    ]


    task = f'select sku,descript from muslin.stock_main where sku like "%-%"'
    result = db.query(task)
    result = list(result.fetchall())

    # Populate sku_dict from the database query result
    sku_dict = {sku.split('-')[0]: description for sku, description in result}

    sku_dict_model = {i: " ".join(description.split(' ')[1:-1]) for i, description in sku_dict.items() if i in idsell}
    df = pd.DataFrame(sku_dict_model.items(), columns=["Value from idsell", "Dict value with idsell"])
    df.to_excel('model.xlsx', index=True)
    print("Excel file 'model.xlsx' created successfully.")

def update_RMA_name():
    task = f'select descript from stock_main'
    result = db.query_custom(task,'muslin')
    result =list(result.fetchall())


    rm_task = f'select descript from RMA_data'
    rm_result = db.query_custom(rm_task,'muslin')
    rm_result =list(rm_result.fetchall())
    rm_result = [i[0].lstrip() for i in rm_result]
    for i in rm_result:
        if 'SN1175-XL' in i[0]:
            print(i[0])

    for i in result:
        sku = i[0]
        if '[' in sku and ']' in sku:
            old_sku = str(sku).split('] ')[1].lstrip()
            if old_sku in rm_result:
                db.query_commit(f'update muslin.RMA_data set descript = "{sku}" where descript = "{old_sku}"')

# checkstock_zort_all('check stock 05092023 final.xlsx','muslin')
def bringback_from_vrich_increase_stock_live(dep,path):
    df = pd.read_excel(path,sheet_name='Sheet2')
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    web.transfer_all_amount_with_condition(0,"W0003","W0001","โยกสต็อกกลับจากไลฟ์")
    web.transfer_all_amount_with_condition(0,"W0003","W0001","โยกสต็อกกลับจากไลฟ์2")
    web.transfer_all_amount_with_condition(0,"W0003","W0001","โยกสต็อกกลับจากไลฟ์3")
    time.sleep(3)
    SKU = web.get_all_sku_by_warehouse('W0001')
    maruay_sku = [i[0] for i in list(db.query_custom('select sku from stock_main','maruay').fetchall())]
    data_list = []
    for i in df.index:
        sku = df.loc[i,'รหัสสต็อก']
        amount = int(df.loc[i,'ขายไป'])
        if sku in SKU and amount > 0 and sku not in maruay_sku:
            data_form = {"sku":sku,"name":sku,"number":amount,"pricepernumber":0,"discount":"0","totalprice":0}
            data_list.append(data_form)

    zort_form = {
    "number":f"ขายใน vrich วันที่ {datetime.date.today()}",
    'customername':"Vrich",
    "orderdate": f"{datetime.date.today()}",
    "amount": 0,
    "warehousecode":"W0001",
    "list":data_list
    }
    web.post_order(zort_form)
    if dep == 'muslin':
        data_list = []
        web = Web("pteRXLvqBNcUXlgIB3RDHxBn3vXoi9cwRp6u/v9M=","GbJK2j7YS5dJtVaBomSsdyenjYuEwdI2A4gLPbKrRAI=","Maruay18.co.th@gmail.com")
        for i in df.index:
            if 'ก' in df.loc[i,'รหัสสินค้า']:
                sku = df.loc[i,'รหัสสต็อก']
                amount = int(df.loc[i,'ขายไป'])
                data_form = {"sku":sku,"name":sku,"number":amount,"pricepernumber":0,"discount":"0","totalprice":0}
                data_list.append(data_form)
        zort_form = {
        "number":f"	VR ก. วันที่ {datetime.date.today()}",
        'customername':"Vrich",
        "orderdate": f"{datetime.date.today()}",
        "amount": 0,
        "warehousecode":"W0001",
        "list":data_list
        }
        web.post_order(zort_form)
# bringback_from_vrich_increase_stock_live('muslin','order1.xlsx')

def transfer_qc():
    path = 'ห้องไลฟ์.xlsx'
    df = pd.read_excel(path)
    payload_list = []

    for i in df.index:
        sku = df.loc[i,'รหัสสินค้า']
        amount = int(df.loc[i,'จำนวน'])
        payload_list.append({'sku':sku,'stock':amount,'cost':0})
        if len(payload_list) > 450:
            web.update_available_stock_list('W0002',payload_list)
            payload_list = []
    web.update_available_stock_list('W0002',payload_list)
    # web.post_transfer('W0001','W0002',f'QC {datetime.date.today()}',payload_list)
# transfer_qc()
def add_korkai_to_zort(path):
    df = pd.read_excel(path)

    column = df.columns
    df.dropna(subset = [column[0]], inplace=True)
    sku = [df.loc[i,column[0]] for i in df.index]
    name = [df.loc[i,column[2]] for i in df.index]
    amount = [df.loc[i,column[4]] for i in df.index]
    cost = [df.loc[i,column[6]] for i in df.index]

    data = list(zip(sku,name,amount,cost))
    data_list,sum_all,refername = [],0,f'korkai {datetime.date.today()}'
    for i in range(len(data)):
        sku,name,amount,price = data[i][0],data[i][1],int(data[i][2]),float(data[i][3])
        data_form = {"sku":sku,"name":name,"number":amount,"pricepernumber":0,"discount":"0","totalprice":0}
        data_list.append(data_form)

    zort_form = {
    "reference":f"{refername}",
    "number": f"{refername}",
    "purchaseorderdate": f"{datetime.date.today()}",
    "amount": 0,
    "warehousecode":"W0003",
    "paymentmethod" : "Cash",
    "paymentamount":0,
    "status" : "Success",
    "list":data_list
    }
    web.post_purchase_order(zort_form)

# add_korkai_to_zort('korkai.xlsx')
# web.transfer_all_amount_with_condition(0,'W0003','W0001')
# transfer_qc()
# web.set_zero_live_warehouse("W0002")
# web.transfer_all_amount_with_condition(4,'W0001','W0003','โยกสต็อกไปไลฟ์ 2')
def extract_digits(input_string):
    # Use regular expression to find all digits in the input string
    digits = re.findall(r'\d+', input_string)
    
    # Combine the digits into a single string
    result = ''.join(digits)
    
    try:
        return int(result)
    except:
        return 0

def export_QC(firstdate, lastdate):
    task = f'select DATE_ADD(date, INTERVAL 7 HOUR),id,image,name from muslin.qc_data where date > "{firstdate}" and date < "{lastdate}"'
    result = db.query(task)
    result = list(result.fetchall())

    workbook = xlsxwriter.Workbook(f'{settings.MEDIA_ROOT}/export_data.xlsx')
    # Create a dictionary to map the sheet names with the respective data
    sheet_data = {}
    sheet_name_list = []
    # Initialize the row index within each sheet
    sheet_row_index = {}

    # Iterate over the results
    for i in range(len(result)):
        # Get the current sheet name
        temp = get_text_after_alphabet(result[i][3].strip())
        first_letter = temp[0]
        if first_letter == 'M' and not str(temp[1]).isdigit():
            sheet_name = "Maruay"
        else:
            second_letter = temp[1]
            if str(second_letter).isdigit():
                second_letter = temp[0]
            sheet_name = f"{second_letter.upper()}"
            if second_letter.upper() in ['B','C','D']:
                sheet_name = 'B,C,D'
            elif second_letter.upper() not in ['A','B','C','D','G']:
                digit = extract_digits(temp.split('-')[0])
                if digit > 1999 and digit < 3000:
                    sheet_name = '2000'
                elif digit > 2999 and digit < 4000:
                    sheet_name = '3000'
                elif digit > 3999 and digit < 5000:
                    sheet_name = '4000'
                else:
                    if second_letter.upper() in ['Y']:
                        sheet_name = "Y"
                    else:
                        sheet_name = "Satin"


        # Check if the sheet_name already exists in the dictionary
        if sheet_name not in sheet_data:
            sheet_data[sheet_name] = []

        # Check if the sheet_row_index exists for the current sheet, initialize if not
        if sheet_name not in sheet_row_index:
            sheet_row_index[sheet_name] = 1

        # Get the current row index and increment it for the next iteration
        row_index = sheet_row_index[sheet_name]
        sheet_row_index[sheet_name] += 20

        # Save the image
        with open(f"{settings.MEDIA_ROOT}/{result[i][1]}.jpg", 'wb') as f:
            f.write(result[i][2])

        # Insert data and images into the worksheet
        if sheet_name not in sheet_name_list:
            worksheet = workbook.add_worksheet(sheet_name)
            worksheet.set_column('A:A', 30)
            sheet_name_list.append(sheet_name)
        else:
            worksheet = workbook.get_worksheet_by_name(sheet_name)

        worksheet.write(f'A{row_index}', str(result[i][0]))
        worksheet.write(f'B{row_index}', result[i][3])
        worksheet.insert_image(f'C{row_index}', f'{settings.MEDIA_ROOT}/{result[i][1]}.jpg', {'x_scale': 0.2, 'y_scale': 0.2})
    # Sort the sheet_name_list alphabetically
    sheet_name_list = sorted(sheet_name_list)

    # Reorder the worksheets based on the sorted sheet_name_list
    workbook.worksheets_objs.sort(key=lambda x: sheet_name_list.index(x.get_name()))

    workbook.close()
# export_QC('2023-10-29','2023-11-12')
def filter_telephone_numbers(file_path):
    filtered_numbers = []

    # Extract the filename and extension
    file_name = os.path.basename(file_path)
    file_name_without_ext = os.path.splitext(file_name)[0]

    # Read the file with the appropriate encoding
    with open(file_path, 'r', encoding='utf-8') as file:
        numbers = file.readlines()

    # Apply regex filters
    for number in numbers:
        # Filter 1: Remove non-numeric characters
        number = re.sub(r'[^0-9]', '', number)

        # Filter 2: Replace numbers starting with 66 with 0
        number = re.sub(r'^66', '', number)

        # Filter 3: Split numbers into two lines if they have 10 digits twice
        number = re.sub(r'(\d{10})(\d{10})', r'\1\n\2', number)

        # Filter 4: Remove lines with numbers that are not exactly 10 digits
        number = re.sub(r'^(?!\d{10}$).*\n', '', number)

        # Add the filtered number to the list
        if number:
            filtered_numbers.append(number)

    # Concatenate the filtered numbers with new lines
    concatenated_numbers = '\n'.join(filtered_numbers)

    # Write the concatenated numbers to a new file
    filtered_file_name = f"filtered_{file_name_without_ext}.txt"
    with open(filtered_file_name, 'w', encoding='utf-8') as filtered_file:
        filtered_file.write(concatenated_numbers)

    return filtered_file_name

class Order:
    def __init__(self, order_number):
        self.number = f"VRM-{order_number}"
        self.list = []

    def add_product(self, sku, name, quantity,pricepernumber):
        product = {"sku": sku, "name": name, "number": quantity,"pricepernumber":pricepernumber,"totalprice": quantity * pricepernumber}
        self.list.append(product)

    def to_dict(self):
        dict_form = {}

        for attribute_name in dir(self):
            if not attribute_name.startswith("__") and not callable(getattr(self, attribute_name)):
                key =attribute_name
                value = getattr(self, attribute_name)
                dict_form[key] = value

        return dict_form

# Function to read Excel file and create orders
def get_vrich_order(file_path):
    df = pd.read_excel(file_path)
    df = df[(~df['เลขพัสดุ'].isnull())]
    df.replace('@#%$-&$#!', '', inplace=True)
    # df.dropna(subset = ['ขนส่ง'], inplace=True)
    orders = {}
    for index, row in df.iterrows():
        order_number = row["เลขที่ Order"]
        # Check if order_number already exists, create a new Order if not
        if order_number not in orders:
            orders[order_number] = Order(order_number)

        # Add product to the respective order
        orders[order_number].add_product(row["รหัสสต็อก"], row["รายละเอียด"], row["จำนวน"],row["ราคา"])
        orders[order_number].address = row["ที่อยู่"]
        orders[order_number].amount = row["รวม"]
        orders[order_number].customeremail = row["ชื่อ Facebook"]
        orders[order_number].customername = row["ชื่อลูกค้า"]
        orders[order_number].customeraddress = row["ที่อยู่"]
        orders[order_number].customerphone = row["โทรศัพท์"]
        orders[order_number].shippingname = row["ชื่อลูกค้า"]
        orders[order_number].shippingamount = row["ค่าส่ง"]
        orders[order_number].vattype = 3
        orders[order_number].shippingvat = 1
        orders[order_number].shippingaddress = row["ที่อยู่"]
        orders[order_number].shippingphone = row["โทรศัพท์"]
        orders[order_number].shippingchannel = 'ไปรษณีย์ไทย'
        orders[order_number].trackingno = row['เลขพัสดุ']
        orders[order_number].paymentmethod = 'โอนผ่านธนาคาร'
        orders[order_number].orderdate = row['วันที่'].strftime("%Y-%m-%d")
        orders[order_number].paymentamount = row["โอนแล้ว"]
        orders[order_number].saleschannel = 'Vrich'
        orders[order_number].warehousecode = 'W0001'
        if orders[order_number].orderdate == '2023-10-16' or orders[order_number].orderdate == '2023-11-15' or orders[order_number].orderdate == '2023-11-16':
            orders[order_number].warehousecode = 'W0002'
        orders[order_number].isCOD = False
        try:
            orders[order_number].paymentdate = row["วัน-เวลาโอน"].strftime("%Y-%m-%d %H:%M")
        except:
            pass
            
        if int(row['โอนแล้ว']) == 0:
            orders[order_number].shippingchannel = 'EMS (COD)'
            orders[order_number].paymentmethod = 'COD'
            orders[order_number].isCOD = True
            orders[order_number].customername += ' COD'
        
        if row["ฝากของ"]:
            orders[order_number].trackingno = ''
        
    return orders

    # Example usage:
    
def note_dobybot(dep,order,note):
    token = {
        'muslin':'8942efad52e0f91ece81c141e10b19753bb91136',
        'maruay':'86c3496d32e2dc406c4d562c40c4a2847785d056',
    }
    paymentstatus = "Paid"
    if order.paymentamount != order.amount:
        paymentstatus = "Pending"

    url = "https://api.dobybot.com/api/picking/orders/sync/"
    dbb_productlist = []
    for i in order.list:
        product = {'sku':i['sku'],'sku_barcod':i['sku'],'name':i['name'],'number':i['number'],'totalprice':i['totalprice'],'pricepernumber':i['pricepernumber']}
        dbb_productlist.append(product)

    payload = [{
    "id": order.number,
	"amount": order.amount,
	"createdatetimeString": f"{order.orderdate} 00:00",
	"customeraddress": order.customeraddress,
	"customername": order.customername,
	"customerphone": order.customerphone,
	"discountamount": "0",
	"isCOD": order.isCOD,
	"list": dbb_productlist,
	"number": order.number,
	"orderdateString": order.orderdate,
	"paymentamount": order.paymentamount,
	"payments": [
		{
			"name": order.paymentmethod,
			"amount": order.paymentamount,
			"paymentdatetimeString": f"{order.orderdate} 00:00"
		}
	],
	"paymentstatus": paymentstatus,
	"saleschannel": "VRich",
	"shippingaddress": order.shippingaddress,
	"shippingamount": order.shippingamount,
	"shippingchannel": order.shippingchannel,
	"shippingname": order.shippingname,
	"shippingphone": order.shippingphone,
	"status": "Pending",
	"trackingno": order.trackingno,
	"vatamount": order.amount * 7 / 100,
	"vattype": 3,
    "integrationName": "VRich",
    "integrationShop": "",
    "remark": note
    }]

    headers = {
    'Authorization': f'Token {token[dep]}',
    'Content-Type': 'application/json'
    }
    print(payload)
    response = requests.post(url, headers=headers, data=json.dumps(payload))
    print(response)
    print(response.text)

def get_note_vrich(dep):
    task = 'select detail from blog where answer = "Dbb"'
    result = db.query_custom(task,dep)
    result = result.fetchall()
    a = {}
    for i in result:
        a[f"VRM-{i[0].split('/')[0].strip()}"] = i[0].split('/')[1]
    return a

def pull_order_vrich_to_zort(dep,file_path,korkai_check):
    start_time = time.time()
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    order_created = []

    # get order from excel file and convert to class Order by order number
    orders_dict = get_vrich_order(file_path)

    # get order that have note in Dobybot while not pull order to zort yet
    order_vrich_note = get_note_vrich(dep)

    # transfer amount back from live warehouse to main warehouse
    web.transfer_all_amount_with_condition(0,"W0003","W0001","โยกสต็อกกลับจากไลฟ์")
    web.transfer_all_amount_with_condition(0,"W0003","W0001","โยกสต็อกกลับจากไลฟ์2")
    web.transfer_all_amount_with_condition(0,"W0003","W0001","โยกสต็อกกลับจากไลฟ์3")
    
    time.sleep(3)

    for i in orders_dict:
        zort_form = orders_dict[i].to_dict()
        temp = ''

        # generate list of order already created
        if not order_created:
            order_created = web.get_order_vrich(orders_dict[i].orderdate)
            
        # if in created_list then pass
        if orders_dict[i].number in order_created:
            continue

        # if its COD order then note in Dobybot
        if orders_dict[i].isCOD:
            temp += f"COD ยอด {orders_dict[i].amount} "
            note_dobybot(dep,orders_dict[i],f"COD ยอด {orders_dict[i].amount}")
                
        # if no trackingno then means deposit order then note in Dobybot            
        if orders_dict[i].trackingno == '':
            temp += f"ออเดอร์ฝากของ "
            print(i,'ฝากของ')
            note_dobybot(dep,orders_dict[i],f"ออเดอร์ฝากของ")

        # if in order_vrich_note means Admin note for some reason then note it with admin's note
        try:
            if orders_dict[i].number in order_vrich_note:
                temp += order_vrich_note[orders_dict[i].number]
                print(temp)
                note_dobybot(dep,orders_dict[i],temp)
        except Exception as e:
            print(f'{e} with order {i}')

        # create an order
        web.post_order(zort_form)

    # ----------------------------------------------------------------------------------------------------------

    print(f"Time taken by to_dict: {time.time() - start_time} seconds")
    web.get_minus_available('W0001')
    web.get_minus_available('W0002')
    if dep == 'muslin' and korkai_check:
        df = pd.read_excel(file_path,sheet_name='Sheet2')
        data_list = []
        web = Web("pteRXLvqBNcUXlgIB3RDHxBn3vXoi9cwRp6u/v9M=","GbJK2j7YS5dJtVaBomSsdyenjYuEwdI2A4gLPbKrRAI=","Maruay18.co.th@gmail.com")
        for i in df.index:
            if 'ก' in df.loc[i,'รหัสสินค้า']:
                sku = df.loc[i,'รหัสสต็อก']
                amount = int(df.loc[i,'ขายไป'])
                data_form = {"sku":sku,"name":sku,"number":amount,"pricepernumber":0,"discount":"0","totalprice":0}
                data_list.append(data_form)
        zort_form = {
        "number":f"	VR ก. วันที่ {datetime.date.today()}",
        'customername':"Vrich",
        "orderdate": f"{datetime.date.today()}",
        "amount": 0,
        "warehousecode":"W0001",
        "list":data_list
        }
        web.post_order(zort_form)

def transfer_all_fromwarehouse_by_path(path):
    df = pd.read_excel(path)
    idsell = [get_idsell(df.loc[i,'sku']) for i in df.index]
    web.transfer_all_amount_with_condition_by_idsell(idsell,'W0001','W0003')

def transfer_all_fromwarehouse_by_idsell(data):
    web.transfer_all_amount_with_condition_by_idsell(data,'W0001','W0003')

def filter_telephone_numbers(dep):
    file_path = os.path.join(settings.MEDIA_ROOT, 'stock', f'tel {datetime.datetime.today().strftime("%Y-%m-%d")} {dep}.txt')
    task = """
    SELECT DISTINCT tel,orderdate
    FROM deli_zort 
    WHERE tel NOT LIKE '%*%' 
    ORDER BY orderdate DESC limit 10500;
    """
    result = list(db.query_custom(task,dep).fetchall())
    with open(file_path, 'w',encoding='utf-8') as file:
        for row in result:
            file.write(str(row[0]) + '\n')

    filtered_numbers = []

    # Extract the filename and extension
    file_name = os.path.basename(file_path)
    file_name_without_ext = os.path.splitext(file_name)[0]

    # Read the file with the appropriate encoding
    with open(file_path, 'r', encoding='utf-8') as file:
        numbers = file.readlines()

    # Apply regex filters
    for number in numbers:
        # Filter 1: Remove non-numeric characters
        number = re.sub(r'[^0-9]', '', number)

        # Filter 2: Replace numbers starting with 66 with 0
        number = re.sub(r'^66', '', number)

        # Filter 3: Split numbers into two lines if they have 10 digits twice
        number = re.sub(r'(\d{10})(\d{10})', r'\1\n\2', number)

        # Filter 4: Remove lines with numbers that are not exactly 10 digits
        number_match = re.search(r'\b\d{10}\b', number)
        if number_match:
            filtered_numbers.append(str(number_match.group()) + '\n')

    # Concatenate the filtered numbers with new lines
    concatenated_numbers = ''.join(filtered_numbers)

    # Write the concatenated numbers to a new file
    filtered_file_name_without_ext = f"filtered_{file_name_without_ext}.txt"
    filtered_file_name = os.path.join(settings.MEDIA_ROOT, 'stock', f'filtered_tel {datetime.datetime.today().strftime("%Y-%m-%d")} {dep}.txt')
    with open(filtered_file_name, 'w', encoding='utf-8') as filtered_file:
        filtered_file.write(concatenated_numbers)

    return filtered_file_name_without_ext

def insert_live_room(path):
    df = pd.read_excel(path)

    # Use iloc to get values from the first and second columns
    sku_values = ','.join([f'("{sku}", {value})' if not pd.isnull(value) else f'("{sku}", null)' for sku, value in zip(df.iloc[:, 0], df.iloc[:, 1])])

    task = f'insert into muslin.live_room values {sku_values}'
    db.query_commit(task)

def print_barcode_from_excel(path):
    df = pd.read_excel(path)
    column = df.columns
    data = []
    for i in df.index:
        for amount in range(int(df.loc[i,column[1]])):
            data.append(df.loc[i,column[0]])
    return data

def setZeroReturnWarehouse(dep):
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    web.set_zero_live_warehouse('W0003')

# transfer_all_fromwarehouse_by_idsell('sku.xlsx')
def update_qty_from_first_file(firstpath,secondpath):
    sku_dict = {}
    df = pd.read_excel(firstpath,skiprows=[0],skipfooter=3)
    for i in df.index:
        sku,amount = df.loc[i,'รหัสสินค้า'],int(df.loc[i,'จำนวนพร้อมขาย'])
        sku_dict[sku] = amount
    try:
        df_platform = pd.read_excel(secondpath)
        for i in df_platform.index:
            if i > 5:
                sku = df_platform.loc[i,'et_title_variation_sku']
                df_platform.loc[i,'et_title_variation_id'] = df_platform.loc[i,'et_title_variation_id']
                df_platform.loc[i,'et_title_product_id'] = df_platform.loc[i,'et_title_product_id']
                if sku in sku_dict and sku_dict[sku] > 0:
                    df_platform.loc[i,'et_title_variation_stock'] = sku_dict[sku]
                else:
                    df_platform.loc[i,'et_title_variation_stock'] = 0
    except:
        df_platform = pd.read_excel(secondpath)
        for i in df_platform.index:
            if i > 4:
                sku = df_platform.loc[i,'seller_sku']
                df_platform.loc[i,'product_id'] = str(df_platform.loc[i,'product_id'])
                df_platform.loc[i,'sku_id'] = str(df_platform.loc[i,'sku_id'])
                if sku in sku_dict and sku_dict[sku] > 0:
                    df_platform.loc[i,'quantity'] = sku_dict[sku]
                else:
                    df_platform.loc[i,'quantity'] = 0
                
    df_platform.to_excel(f'{settings.MEDIA_ROOT}/stock/final.xlsx',index=False)
    return f'media/stock/final.xlsx'

# update_qty_from_first_file('first.xlsx','second.xlsx')
def update_name_by_replace(dep,sku_code,replace_word,replace_by):
    web = Web(get_api_register(dep,'apikey'),get_api_register(dep,'apisecret'),get_api_register(dep,'storename'))
    task = f'select sku from stock where sku like "%{sku_code}%"'
    result = db.query_custom(task,dep)

    result = result.fetchall()
    for i in result:
        web.update_name(dep,i[0],replace_word,replace_by)
